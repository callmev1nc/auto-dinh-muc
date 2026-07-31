import openpyxl, sys, io

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"E:\Work\Thai-work\auto-dinh-muc\templates\Định mức - Bao giấy (KP - in offset).xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

targets = ["Tráng", "Dán", "Thổi", "May", "In"]

for name in targets:
    if name not in wb.sheetnames:
        print(f"\n=== {name} === [NOT FOUND]")
        continue
    ws = wb[name]
    print(f"\n=== {name} ===")

    # --- C5:C9 header ---
    print("  C5..C9 (header):")
    for r in range(5, 10):
        v = ws.cell(r, 3).value
        print(f"    C{r} = {v!r}")

    # --- M6:M14 info table ---
    print("  M6..M14 (info table):")
    for r in range(6, 15):
        v = ws.cell(r, 13).value
        vn = ws.cell(r, 14).value
        merged_note = ""
        if vn is not None:
            merged_note = f"  [N{r}={vn!r}]"
        print(f"    M{r} = {v!r}{merged_note}")

    # --- detect merged cells touching M column ---
    merged = []
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= 13 <= mr.max_col:
            merged.append(str(mr))
    if merged:
        print(f"  Merged ranges involving M column: {merged}")
    else:
        print("  No merged cells in M column")

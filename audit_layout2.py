import openpyxl, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"E:\Work\Thai-work\auto-dinh-muc\templates\Định mức - Bao giấy (KP - in offset).xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

targets = ["Tráng", "Dán", "Thổi", "May", "In"]

for name in targets:
    if name not in wb.sheetnames:
        print(f"\n=== {name} === [NOT FOUND]")
        continue
    ws = wb[name]
    print(f"\n=== {name} ===  (dims={ws.dimensions})")

    # --- C5:C9 header ---
    print("  C5..C9 (header):")
    for r in range(5, 10):
        v = ws.cell(r, 3).value
        print(f"    C{r} = {v!r}")

    # Determine M range: look at all merged ranges involving M
    m_ranges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= 13 <= mr.max_col:
            m_ranges.append((mr.min_row, mr.max_row))
    if m_ranges:
        min_m_row = min(r[0] for r in m_ranges)
        max_m_row = max(r[1] for r in m_ranges)
        print(f"  M-column merged cell row span: {min_m_row}..{max_m_row}")
    else:
        min_m_row, max_m_row = 6, 14

    # Dump M column across the full merged span
    print(f"  M{min_m_row}..M{max_m_row} (info table):")
    for r in range(min_m_row, max_m_row + 1):
        v = ws.cell(r, 13).value
        vn = ws.cell(r, 14).value
        note = f"  [N{r}={vn!r}]" if vn is not None else ""
        print(f"    M{r} = {v!r}{note}")

    # List all merged ranges involving M
    merged = sorted(str(mr) for mr in ws.merged_cells.ranges if mr.min_col <= 13 <= mr.max_col)
    print(f"  Merged ranges in M column: {merged}")
    
    # Also dump C column for In to see header structure
    if name == "In":
        print("  C1..C50 dump for In:")
        for r in range(1, 51):
            v = ws.cell(r, 3).value
            if v is not None:
                print(f"    C{r} = {v!r}")

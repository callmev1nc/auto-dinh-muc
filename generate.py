#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate.py — Customer order -> Định mức + YCSX Excel files.

ARCHITECTURE (hybrid, byte-faithful):
  * Clone a template .xlsx and fill ONLY the input cells (cell_map.json).
  * Use xlsxpatch.XlsxPatch so embedded images/drawings/styles are preserved
    (openpyxl would drop them). Output opens identically to the examples.
  * The ONE interactive question is "số màu in?" (number of print colors) —
    every other value is parsed from the order or computed from formulas.json.

USAGE:
  python generate.py --order order.json            # fill from a JSON order
  python generate.py --ycsx path/to/YCSX.xlsx      # parse fields from a YCSX file
  python generate.py --sample                      # run the built-in 40KG sample
  python generate.py --order order.json --colors 3 # pre-answer the only question
"""
from __future__ import annotations
import argparse, json, os, sys, datetime, re, copy, zipfile
from xlsxpatch import XlsxPatch
import nvl_lookup as nvl
import tsvh_lookup as tsvh

try:  # Windows console defaults to cp1252; force utf-8 so Vietnamese prints survive
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass
try:
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")
TEMPLATES = os.path.join(HERE, "templates")
OUTDIR = os.path.join(HERE, "output")
NVL_PATH = os.path.join(DATA, "Nguyên vật liệu.xlsx")


def template_path(key):
    return os.path.join(TEMPLATES, {
        "opp": "Định mức - OPP (Bao BOPP in ống đồng).xlsx",
        "paper_kp": "Định mức - Bao giấy (KP - in offset).xlsx",
        "ycsx": "YCSX.xlsx",
    }[key])


def load_json(name):
    with open(os.path.join(DATA, name), encoding="utf-8") as f:
        return json.load(f)


CONST = load_json("constants.json")
CELLMAP = load_json("cell_map.json")
RULES = load_json("customer_rules.json")
BAG = load_json("bag_type_map.json")
DIMS = load_json("standard_dims.json")

# Máy in tối đa 6 màu (khớp constants.phe_chong_mau_m chỉ có key 1..6).
SO_MAU_IN_MAX = int(CONST.get("so_mau_in_max", 6))


# ---------------------------------------------------------------- parse order
def parse_ycsx(path):
    """Parse a YCSX .xlsx (read-only): header + ALL product line items.

    Merge-aware: a cell covered by a merged range returns that range's top-left
    value, so a shared spec such as F13:H15 applies to every product row beneath
    it. Returns an order dict with header fields, a ``products`` list (one entry
    per product row that carries a qty), and top-level single-product fields set
    to the first product (backward-compat with --order / --sample callers).

    Formula-tolerant: loads the workbook twice — ``data_only=True`` for cached
    values and ``data_only=False`` for raw formula strings.  When ``cellv``
    finds a formula that has no cached value (e.g. file saved by a non-Excel
    tool), it returns ``None`` so the caller can distinguish "empty" from
    "uncached formula".
    """
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries, coordinate_to_tuple, get_column_letter
    wb_data = load_workbook(path, data_only=True)
    wb_fml = load_workbook(path, data_only=False)
    ws = wb_data[wb_data.sheetnames[0]]
    ws_fml = wb_fml[wb_fml.sheetnames[0]]

    def _is_formula(coord):
        v = ws_fml[coord].value
        return isinstance(v, str) and v.startswith("=")

    def cellv(coord):
        """Return value or None.  Returns None explicitly when the cell holds a
        formula that has no cached value (so callers don't mistake it for
        empty or fall through to a stale merged-cell neighbour)."""
        val = ws[coord].value
        if val is not None:
            return val
        if _is_formula(coord):
            return None
        cr, cc = coordinate_to_tuple(coord)
        for mr in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(mr))
            if min_row <= cr <= max_row and min_col <= cc <= max_col:
                merged_val = ws.cell(row=min_row, column=min_col).value
                if merged_val is not None:
                    return merged_val
                merged_coord = ws_fml.cell(row=min_row, column=min_col).coordinate
                if _is_formula(merged_coord):
                    return None
                return merged_val
        return None

    # header labels live in column B as "N. Label: value"
    g = {c.coordinate: c.value.strip()
         for row in ws.iter_rows(values_only=False)
         for c in row if isinstance(c.value, str)}

    def bfind(label):
        """(coord, value) của ô đầu tiên mang nhãn header + danh sách ô TRÙNG.

        Trả cả toạ độ để fill_ycsx ghi ĐÚNG ô đã đọc ra, thay vì ghi cứng B6:B10
        (đơn Hùng Duy 25S03IKH00502000046: layout lệch dòng nên bản ghi cứng tạo
        thêm một dòng "5. Số đơn hàng" thứ hai ngay trên bảng sản phẩm).
        """
        hits = [(coord, val) for coord, val in g.items() if val and val.startswith(label)]
        if not hits:
            return None, "", []
        coord, val = hits[0]
        value = val.split(":", 1)[-1].strip() if ":" in val else ""
        return coord, value, [c for c, _ in hits[1:]]

    HEADER_LABELS = {
        "ngay_yc": "1. Ngày yêu cầu",
        "customer": "2. Khách hàng",
        "address": "3. Địa chỉ",
        "customer_code": "4. Mã khách hàng",
        "order_id": "5. Số đơn hàng",
    }
    header_cells, dup_cells = {}, []
    order = {}
    for field, label in HEADER_LABELS.items():
        coord, value, dups = bfind(label)
        order[field] = value
        if coord:
            header_cells[field] = {"ref": coord, "label": label, "text": g[coord]}
        dup_cells.extend(dups)
    if not order["customer"]:
        order["customer"] = g.get("B7", "")

    # --- locate the product-table header row + columns by header labels --------
    # Portable across customers: a sale YCSX may start the table on a different
    # row or put 'Số lượng' in a different column than the 4 Oranges template.
    # The header row is the first row that has a 'Số lượng' column AND a
    # 'Mã sản phẩm'/'Tên sản phẩm' column. Merge-aware cellv() still resolves
    # shared specs (e.g. F13:H15) for every product row beneath the header.
    QTY_KEYS = ("số lượng", "so luong", "số lương", "qty")
    CODE_KEYS = ("mã sản phẩm", "ma san pham", "mã sp", "mã hàng")
    NAME_KEYS = ("tên sản phẩm", "ten san pham", "tên hàng", "mặt hàng")
    MACODE_KEYS = ("mã code", "ma code", "mtls")
    SPEC_KEYS = ("chi tiết kỹ thuật", "chi tiet ky thuat", "thông số", "quy cách")
    DVT_KEYS = ("đvt", "dvt", "đơn vị", "don vi")

    def _col_with(row, keys):
        for c in range(1, 24):
            v = ws.cell(row, c).value
            if isinstance(v, str) and any(k in v.strip().lower() for k in keys):
                return get_column_letter(c)
        return None

    header_row, qty_col = 12, "J"  # historical fallback layout
    for r in range(1, 32):
        qc = _col_with(r, QTY_KEYS)
        if qc and (_col_with(r, CODE_KEYS) or _col_with(r, NAME_KEYS)):
            header_row, qty_col = r, qc
            break
    code_col = _col_with(header_row, CODE_KEYS) or "C"
    name_col = _col_with(header_row, NAME_KEYS) or "D"
    ma_col = _col_with(header_row, MACODE_KEYS) or "E"
    spec_col = _col_with(header_row, SPEC_KEYS) or "F"
    # ĐVT dò theo nhãn — KHÔNG mặc định về "I": khách nào không có cột "Mã code"
    # thì mọi cột lệch đi 1 và ghi cứng "Cái" vào I sẽ đè lên cột khác.
    dvt_col = _col_with(header_row, DVT_KEYS)

    # product rows: any row under the header with a numeric qty + a code/name.
    # Stage-marker rows (MÀNH/IN/TRÁNG/...) carry no qty -> skipped.
    products = []
    for r in range(header_row + 1, header_row + 41):
        code, name = cellv(f"{code_col}{r}"), cellv(f"{name_col}{r}")
        qty = _parse_qty(cellv(f"{qty_col}{r}"))
        if qty is None or qty <= 0 or not (code or name):
            continue
        products.append({
            "product_code": str(code).strip() if code else "",
            "product_name": str(name).strip() if name else "",
            "ma_code": str(cellv(f"{ma_col}{r}") or "").strip(),
            "qty": qty,
            "spec": str(cellv(f"{spec_col}{r}") or "").strip().strip('"'),
        })

    if not products:
        # surface exactly what the số lượng column contained so the cause is
        # obvious (wrong column, blank SL, or SL stored as an uncached formula).
        seen = [f"{qty_col}{r}={cellv(f'{qty_col}{r}')!r}"
                for r in range(header_row + 1, header_row + 16)
                if cellv(f"{qty_col}{r}") not in (None, "")]
        raise InputValidationError(
            f"Không đọc được dòng sản phẩm nào từ phiếu YCSX (header dòng "
            f"{header_row}, cột số lượng = {qty_col}). Ô {qty_col} thấy: "
            f"{', '.join(seen) or '(rỗng)'}. → Kiểm tra cột 'Số lượng' trong "
            f"phiếu YCSX có điền số > 0 cho từng mặt hàng; nếu SL là công thức "
            f"Excel, hãy copy → paste-as-value trước khi upload."
        )
    order["products"] = products
    # layout anchor used to re-fill the SAME file when the input is a YCSX
    order["ycsx_header_row"] = header_row
    order["ycsx_cols"] = {
        "code": code_col, "name": name_col, "ma": ma_col, "spec": spec_col,
        "qty": qty_col, "dvt": dvt_col,
    }
    order["ycsx_header_cells"] = header_cells
    # ô TRÙNG nhãn header — chỉ tính phần NẰM TRÊN bảng sản phẩm để không bao giờ
    # chạm vào dòng sản phẩm / ghi chú giao hàng.
    order["ycsx_dup_header_cells"] = [
        ref for ref in dup_cells if coordinate_to_tuple(ref)[0] < header_row]
    # snapshot để fill_ycsx so sánh trước khi ghi (chỉ ghi ô nào THỰC SỰ khác)
    order["ycsx_snapshot"] = {
        c.coordinate: c.value
        for row in ws.iter_rows(min_row=1, max_row=header_row + 40, max_col=13)
        for c in row if c.value is not None}

    # top-level single-product fields = first product (backward compat)
    p0 = products[0]
    order.update({
        "product_code": p0["product_code"], "product_name": p0["product_name"],
        "ma_code": p0["ma_code"], "qty": p0["qty"], "spec": p0["spec"],
    })
    order.update(parse_spec(p0["spec"]))
    if not order.get("inner_bag_weight_kg"):
        ibw = _extract_pe_liner_weight(p0.get("product_name", ""))
        if ibw is not None:
            order["inner_bag_weight_kg"] = ibw
    return order


# Dấu nhân trong kích thước: sale viết "x", "X", "×" hoặc "*" (đơn Hùng Duy
# 25S03IKH00502000046 ghi "(57*110cm)" → trước đây không parse được, sheet Thổi 2
# bị trống kích thước túi lồng). Character class, KHÔNG phải group — các regex
# dùng nó đang đánh số group theo vị trí.
_MUL = r"[xX×*]"


def parse_spec(spec):
    """Extract dimensions / structure / bag hints from the Chi tiết kỹ thuật text."""
    out = {}
    if not spec:
        return out
    # dimensions like "(42+8) cm x 82cm" (optional parens) or "50x92 cm".
    # Dấu nhân: khách hàng viết cả "x", "X", "×" và "*" (đơn Hùng Duy: "(57*110cm)").
    m = re.search(r"\(?\s*(\d+)\s*\+\s*(\d+)\s*\)?\s*(?:cm\s*)?" + _MUL + r"\s*(\d+(?:\.\d+)?)\s*cm", spec)
    if m:
        out["width_cm"] = int(m.group(1)); out["gusset_cm"] = int(m.group(2))
        out["width_plus_gusset_m"] = (int(m.group(1)) + int(m.group(2))) / 100
        out["bag_length_m"] = float(m.group(3)) / 100
    else:
        m = re.search(r"(\d+(?:\.\d+)?)\s*" + _MUL + r"\s*(\d+(?:\.\d+)?)\s*cm", spec)
        if m:
            out["width_plus_gusset_m"] = float(m.group(1)) / 100
            out["bag_length_m"] = float(m.group(2)) / 100
    ibw = _extract_pe_liner_weight(spec)
    if ibw is not None:
        out["inner_bag_weight_kg"] = ibw
    tl = _parse_tui_long(spec)
    if tl:
        out.update(tl)
    return out


def _parse_tui_long(spec):
    """Parse Quy cách lồng túi PE → {tui_long_loai, tui_long_rong_cm, tui_long_dai_cm,
    tui_long_kg, tui_long_quy_cach_day}. Returns {} when the spec has no liner info."""
    if not spec or not any(k in spec.lower() for k in _LINER_KW):
        return {}
    if any(k in spec.lower() for k in _LINER_NO_KW):
        return {}
    out = {}
    loai = "thường"
    m_loai = re.search(r"(?:túi (?:lồng )?pe|pe)\s*(rin|thường)", spec, re.IGNORECASE)
    if m_loai:
        loai = m_loai.group(1).strip().lower()
    out["tui_long_loai"] = "rin" if loai == "rin" else "thường"
    # Kích thước túi lồng — chỉ tìm trong mệnh đề nói về túi lồng, nếu không
    # regex sẽ bắt nhầm kích thước BAO ở dòng "1.Kích thước: (45+10) x 89cm".
    seg = _tui_long_segment(spec)
    m_dim = re.search(r"(\d+)\s*" + _MUL + r"\s*(\d+)\s*cm", seg, re.IGNORECASE)
    if not m_dim:
        # nhiều YCSX ghi thiếu đơn vị: "túi PE rin 50x92 (20gr) (LTMS)"
        m_dim = re.search(r"(\d{2,3})\s*" + _MUL + r"\s*(\d{2,3})\s*[( ]", seg)
    if m_dim:
        out["tui_long_rong_cm"] = int(m_dim.group(1))
        out["tui_long_dai_cm"] = int(m_dim.group(2))
    m_qc = re.search(r"\((LTMS|MTLS)\)", spec, re.IGNORECASE)
    if not m_qc:
        m_qc = re.search(r"(LTMS|MTLS)", spec, re.IGNORECASE)
    if m_qc:
        out["tui_long_quy_cach_day"] = m_qc.group(1).upper()
    else:
        # YCSX cũng hay ghi bằng chữ: "(túi lồng may không dính đáy)".
        # Xác nhận 2026-08-12: may KHÔNG dính đáy = đáy ngắn, may dính đáy = đáy dài.
        m_day = re.search(r"may\s+(không\s+)?dính\s+đáy", spec, re.IGNORECASE)
        if m_day:
            out["tui_long_quy_cach_day"] = ("may không dính đáy" if m_day.group(1)
                                            else "may dính đáy")
    return out


def _tui_long_segment(spec):
    """Đoạn text nói về túi lồng (từ từ khoá lồng túi tới hết dòng/mục kế tiếp).

    Cần thiết vì spec chứa NHIỀU cặp số dạng "A x B cm" (kích thước bao ở mục 1,
    kích thước túi lồng ở mục 5) — tìm trên cả spec sẽ lấy nhầm số đầu tiên.
    """
    low = spec.lower()
    idx = min((low.find(k) for k in _LINER_KW if low.find(k) >= 0), default=-1)
    if idx < 0:
        return spec
    seg = spec[idx:]
    # cắt ở đầu mục kế tiếp ("6. Quy cách đóng gói: ...") nếu có
    m_next = re.search(r"\n\s*\d+\s*[.)]", seg)
    return seg[:m_next.start()] if m_next else seg


def _parse_nep(spec, order=None):
    """Parse quy cách nẹp từ 'Chi tiết kỹ thuật' → list[{loai, rong_cm, tokens}].

    Một đơn thường dùng NHIỀU nẹp, ví dụ YCSX Hùng Duy:
        "May nẹp đáy bao (Nẹp KP nhật vàng), dán nẹp giấy đầu còn lại nẹp 10cm
         - dán nẹp giấy về cùng 1 bên mặt sau."
    → nẹp KP vàng Nhật (may) + nẹp giấy 10cm (dán).

    loai: "KP" | "GIAY" | "OPP" | "PP".  tokens: các từ màu/xuất xứ lấy ngay
    trong cụm đó, dùng để tra Nguyên vật liệu.xlsx (không phụ thuộc thứ tự từ).
    Nẹp giấy không ghi màu/xuất xứ → thừa hưởng từ giấy thân bao (quyết định
    2026-08-12: "nẹp giấy chọn theo giấy thân bao").
    Trả [] khi spec không nói gì về nẹp.
    """
    if not spec:
        return []
    order = order or {}
    defaults = CONST["may"]["nep_rong_cm_default"]
    # Chiều rộng ghi rời, không gắn với loại nào: "... còn lại nẹp 10cm"
    loose_w = None
    for m in re.finditer(r"nẹp\s+(\d+(?:[.,]\d+)?)\s*cm", spec, re.IGNORECASE):
        loose_w = float(m.group(1).replace(",", "."))

    found = {}
    for m in re.finditer(r"nẹp\s+(kp|giấy|giay|opp|pp)\b", spec, re.IGNORECASE):
        raw = m.group(1).lower()
        loai = {"kp": "KP", "giấy": "GIAY", "giay": "GIAY", "opp": "OPP", "pp": "PP"}[raw]
        # mệnh đề của cụm này: tới dấu ngắt câu gần nhất
        tail = re.split(r"[,;)\n\-–]", spec[m.end():], maxsplit=1)[0]
        m_w = re.search(r"(\d+(?:[.,]\d+)?)\s*cm", tail)
        rong_cm = float(m_w.group(1).replace(",", ".")) if m_w else None
        tokens = [t for t in re.findall(r"[A-Za-zÀ-ỹ]+", tail) if t.lower() in _NEP_TOKENS]
        prev = found.get(loai)
        if prev is None:
            found[loai] = {"loai": loai, "rong_cm": rong_cm, "tokens": tokens}
        else:  # cụm sau bổ sung thông tin cho cụm trước (KP/giấy nhắc nhiều lần)
            if prev["rong_cm"] is None:
                prev["rong_cm"] = rong_cm
            for t in tokens:
                if t not in prev["tokens"]:
                    prev["tokens"].append(t)

    nep = []
    for loai, item in found.items():
        if item["rong_cm"] is None and loai == "GIAY" and loose_w is not None:
            item["rong_cm"] = loose_w      # "dán nẹp giấy ... nẹp 10cm"
        if item["rong_cm"] is None and len(found) == 1 and loose_w is not None:
            item["rong_cm"] = loose_w
        if item["rong_cm"] is None:
            item["rong_cm"] = defaults.get(loai, 6)
            item["rong_cm_default"] = True
        if loai == "GIAY" and not item["tokens"]:
            item["tokens"] = [str(order.get("giay_kraft_mau", "vàng")),
                              str(order.get("giay_kraft_xuatxu", "Nhật"))]
            item["tokens_from_bao"] = True
        nep.append(item)
    return nep


# Từ màu/xuất xứ dùng để nhận diện đúng dòng nẹp trong Nguyên vật liệu.xlsx.
# CHÚ Ý: danh mục gốc viết sai chính tả "Karft" nên KHÔNG có "kraft" ở đây.
_NEP_TOKENS = {"vàng", "trắng", "xám", "nhật", "việt", "vn", "thái", "bãi", "bằng",
               "mờ", "bóng", "xanh", "đỏ", "tím", "cam", "decal"}


def _detect_has_print(order, so_mau_in=None):
    """A bag is printed unless the order/spec/product name says 'không in'.
    so_mau_in=0 overrides everything: 0 means 'không in' (user's explicit choice)."""
    if so_mau_in == 0:
        return False
    if order.get("has_print") is not None:
        return bool(order.get("has_print"))
    text = (" ".join([
        str(order.get("spec", "")), str(order.get("product_name", "")),
    ])).lower()
    return not any(k in text for k in ("không in", "khong in", "không in ấn", "khong in an", "không in hình"))


def _has_tui_long(order):
    if order.get("has_tui_long") is not None:
        return bool(order.get("has_tui_long"))
    text = (" ".join([str(order.get("spec", "")), str(order.get("product_name", ""))])).lower()
    if any(k in text for k in _LINER_NO_KW):
        return False
    if float(order.get("inner_bag_weight_kg") or 0) > 0:
        return True
    return any(k in text for k in _LINER_KW)


def _gusset_m(order, W):
    """Gusset (hông) in metres. Prefer explicit width_cm/gusset_cm; else derive from
    width_plus_gusset minus a separately-given width."""
    if order.get("gusset_cm"):
        return float(order["gusset_cm"]) / 100
    if order.get("width_cm") and order.get("width_plus_gusset_m"):
        return float(order["width_plus_gusset_m"]) - float(order["width_cm"]) / 100
    return 0.0


def _to_num(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except Exception:
        try:
            return float(v)
        except Exception:
            return None


def _same_cell(old, new):
    """True khi ghi `new` vào ô đang chứa `old` sẽ không thay đổi gì thực chất.

    So khớp lỏng vì Excel lưu số dưới dạng chuỗi tuỳ file ('5000' vs 5000).
    """
    if old is None or old == "":
        return new is None or new == ""
    if new is None or new == "":
        return False
    on, nn = _to_num(old), _to_num(new)
    if on is not None and nn is not None:
        return float(on) == float(nn)
    return str(old).strip() == str(new).strip()


_LINER_KW = ("lồng túi", "túi lồng", "pe thường", "pe rin", "pe lồng")
_LINER_NO_KW = ("không lồng túi", "khong long tui", "không lồng", "khong long",
                "không lót túi", "khong lot tui", "không túi lồng", "khong tui long")


def _extract_pe_liner_weight(text):
    if not text:
        return None
    t = text.lower()
    if not any(k in t for k in _LINER_KW):
        return None
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:gr|gram|grams|g)\b", t)
    if m:
        grams = float(m.group(1).replace(",", "."))
        return grams / 1000
    return None


def _parse_qty(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(" ", "")
    dots = s.count(".")
    commas = s.count(",")
    if dots == 0 and commas == 0:
        try:
            return int(s) if s.isdigit() else float(s)
        except ValueError:
            return None
    if dots > 0 and commas > 0:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif commas > 0:
        parts = s.split(",")
        if commas == 1 and len(parts[-1]) <= 2:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif dots > 0:
        parts = s.split(".")
        if dots == 1 and len(parts[-1]) <= 2:
            pass
        else:
            s = s.replace(".", "")
    try:
        return int(s) if "." not in s else float(s)
    except ValueError:
        return None


# ------------------------------------------------------------ detect bag family
def detect_family(order):
    text = " ".join([
        str(order.get("customer", "")),
        str(order.get("product_name", "")),
        str(order.get("spec", "")),
    ]).lower()
    opp = BAG["families"]["opp"]
    kp = BAG["families"]["paper_kp"]
    if any(k.lower() in text for k in opp["name_keywords"] + opp["spec_keywords"]):
        return "opp"
    if any(k.lower() in text for k in kp["name_keywords"] + kp["spec_keywords"]):
        return "paper_kp"
    # fallback by weight token
    mw = re.search(r"(\d{2,3})\s*kgs?", text)
    if mw:
        w = int(mw.group(1))
        return "opp" if w >= 35 else "paper_kp"
    raise InputValidationError("Cannot detect bag family from order. Set order['bag_family'] manually.")


def _bao_kien(bao_type, length_cm):
    """Số bao/kiện từ bảng TSVH ĐÓNG GÓI (TSDGO) theo loại bao + chiều dài (cm)."""
    tiers = CONST["dong_goi_bao_kien"].get(bao_type, [])
    for tier in tiers:
        if length_cm >= tier["len_cm_min"]:
            return tier["bao_kien"]
    return None


# ----------------------------------------------------------- customer overrides
def apply_rules(order, family):
    text = (str(order.get("customer", "")) + " " + str(order.get("spec", ""))).lower()
    ov = {}
    for r in RULES["rules"]:
        m = r["match"]
        if "customer_contains_any" in m and any(k.lower() in text for k in m["customer_contains_any"]):
            ov.update(r["effect"])
        if "spec_contains_any" in m and any(k.lower() in text for k in m["spec_contains_any"]):
            ov.update(r["effect"])
        if m.get("bag_family") == family:
            ov.update(r["effect"])
    return ov


# ---------------------------------------------------------------- validation gate
class InputValidationError(ValueError):
    pass


def validate_inputs(order, family, so_mau_in):
    errors = []
    if float(order.get("qty") or 0) <= 0:
        errors.append(
            "- SL (cột 'Số lượng' trong phiếu YCSX) phải > 0. Nếu cột này là "
            "công thức Excel, hãy copy → Paste Values trước khi upload."
        )
    if float(order.get("bag_length_m") or 0) <= 0:
        errors.append("- Kích thước dài (bag_length_m) không hợp lệ — kiểm tra dòng 'Kích thước' trong spec")
    if float(order.get("width_plus_gusset_m") or 0) <= 0:
        errors.append("- Kích thước ngang (width_plus_gusset_m) không hợp lệ — kiểm tra dòng 'Kích thước' trong spec")
    has_print = _detect_has_print(order, so_mau_in)
    if so_mau_in < 0:
        errors.append("- Số màu in (so_mau_in) không được âm")
    if so_mau_in > SO_MAU_IN_MAX:
        errors.append(f"- Số màu in (so_mau_in) tối đa là {SO_MAU_IN_MAX} — máy in "
                      f"chỉ chạy được {SO_MAU_IN_MAX} màu")
    has_liner = _has_tui_long(order)
    if has_liner and float(order.get("inner_bag_weight_kg") or 0) <= 0:
        errors.append("- inner_bag_weight_kg (Quy cách lồng túi PE) không hợp lệ — kiểm tra dòng 'Quy cách lồng túi PE' trong spec")
    if family == "opp" and has_print:
        km = order.get("kho_mang") or _dims_lookup(order, "mang_in_cm")
        if not km:
            errors.append("- Khổ màng (kho_mang) phải xác định và > 0 đối với bao OPP — kiểm tra kích thước trong spec")
    if errors:
        raise InputValidationError("Thiếu dữ liệu đầu vào — vui lòng kiểm tra:\n" + "\n".join(errors))


# -------------------------------------------------------------------- compute
def compute(order, family, so_mau_in):
    ov = apply_rules(order, family)
    qty = float(order.get("qty") or 0)
    L = float(order.get("bag_length_m") or 0)
    W = float(order.get("width_plus_gusset_m") or 0)
    tol = float(order.get("tolerance", CONST["dung_sai_default"]))
    bao_type = "BOPP" if family == "opp" else "KP"
    # Tỷ lệ taical/hạt màu theo khách hàng: đơn thường 10%/5%; 4 Oranges & Thanh Phụng 7%/8%.
    taical_share = float(ov.get("trang_taical_share", CONST["trang"]["taical_share_default"]))
    hat_mau_share = float(ov.get("trang_hat_mau_share", CONST["trang"]["hat_mau_share_default"]))
    warnings = []
    items = nvl.load_nvl_list(NVL_PATH) if os.path.isfile(NVL_PATH) else []

    has_print = _detect_has_print(order, so_mau_in)
    has_tui_long = _has_tui_long(order)
    ibw = float(order.get("inner_bag_weight_kg") or 0)
    if not has_tui_long:
        ibw = 0.0

    gusset_m = _gusset_m(order, W)
    width_m = max(W - gusset_m, 0.0)

    kho_manh = round(W * 2 + 0.06, 3)
    kho_mang = round(W * 2 + 0.04, 3)
    kho_giay = round(W * 2 + 0.02, 3)
    kho_thanh_pham = round(W * 2 + 0.04, 3)

    # ---- Số lượng in thực tế (In!M9) & downstream Tráng!M9 & thành phẩm (G29)
    # Confirmed rule: BOPP applies tolerance on metres; KP does NOT (tolerance
    # only on thành phẩm). Phế chồng màu is a BOPP-with-print-only addition.
    phe = 0
    if has_print and bao_type == "BOPP":
        phe = CONST["phe_chong_mau_m"].get(str(so_mau_in), 500)
    if bao_type == "BOPP":
        sl_in = qty * L * (1 + tol) + phe
        trang_sl = qty * L * (1 + tol)
    else:
        sl_in = qty * L
        trang_sl = qty * L
    sl_in = round(float(sl_in), 1)
    trang_sl = round(float(trang_sl), 4)
    thanh_pham = int(round(qty)) if bao_type == "BOPP" else int(round(qty * (1 + tol)))

    # ---- NVL lookups (name/code), never guess
    def _exact(keywords):
        if not items:
            return None
        found = nvl.find_exact_name(items, keywords)
        return found[0] if found else None

    manh_dl = int(ov.get("dinh_luong_manh_trang_gm2", CONST["dinh_luong_manh_trang_gm2"]))
    kho_manh_mm = int(round(kho_manh * 1000))
    manh, manh_cands = (None, [])
    if items:
        manh, manh_cands = nvl.find_manh_trang(items, kho_manh_mm, manh_dl)
    if not manh:
        warnings.append(f"Không tìm thấy Mành trắng K{kho_manh_mm} ĐL{manh_dl} trong "
                        f"Nguyên vật liệu.xlsx. Ứng viên gần đúng: {manh_cands}. Cần hỏi lại.")
    f801c = _exact(["Hạt nhựa nguyên sinh F801C"])
    taical = _exact(["Hạt phụ gia EFPP 105BC"])
    vistamaxx = _exact(["Hạt nhựa Vistamax 6202"])

    mang = giay = None
    loai_mang = order.get("mang_opp_loai", "mờ")
    dl_mang = 0.01638 if str(loai_mang).strip() == "bóng" else 0.01584
    if bao_type == "BOPP":
        if items:
            mang, mang_cands = nvl.find_mang_opp(items, loai_mang, int(round(kho_mang * 1000)))
        if not mang:
            warnings.append(f"Không tìm thấy Màng BOPP {loai_mang} khổ {int(round(kho_mang * 1000))}mm "
                            f"trong Nguyên vật liệu.xlsx. Ứng viên gần đúng: "
                            f"{mang_cands if items else '(chưa có file NVL)'}. Cần hỏi lại.")
    else:
        if items:
            giay, giay_cands = nvl.find_giay_kraft(
                items, order.get("giay_kraft_mau", "vàng"),
                order.get("giay_kraft_xuatxu", "Nhật"),
                int(round(kho_giay * 1000)), 70,
                hang_uu_tien=order.get("giay_kraft_hang", "TAIKO"))
        if not giay:
            warnings.append(f"Không tìm thấy Giấy Kraft khổ {int(round(kho_giay * 1000))}mm trong "
                            f"Nguyên vật liệu.xlsx. Ứng viên gần đúng: "
                            f"{giay_cands if items else '(chưa có file NVL)'}. Cần hỏi lại.")

    # ---- In sheet materials
    mang_bopp_kg = round(sl_in * kho_mang * dl_mang, 4) if bao_type == "BOPP" else 0.0
    dung_moai_opp_kg = round(CONST["in_opp"]["dung_moai_opp_coeff"] * sl_in
                             * CONST["in_opp"]["dung_moai_opp_share"], 4)
    dung_moai_ea_kg = round(CONST["in_opp"]["dung_moai_opp_coeff"] * sl_in
                            * CONST["in_opp"]["dung_moai_ea_share"], 4)
    giay_kraft_kg = round(trang_sl * kho_giay * (CONST["dinh_luong_giay_kraft_gm2"] / 1000), 4)

    # ---- Tráng stage
    trang_g17 = round(trang_sl * kho_manh * (manh_dl / 1000), 4)
    t = CONST["trang"]
    if bao_type == "BOPP":
        trang_g18 = round(trang_sl * kho_manh * t["f801c_dinh_luong"] * t["f801c_share_opp"], 4)
        trang_g19 = round(trang_sl * kho_manh * t["taical_dinh_luong"] * taical_share, 4)
        trang_g20 = round(trang_sl * kho_manh * t["hat_mau_dinh_luong"] * hat_mau_share, 4)
    else:
        trang_g18 = round(trang_sl * kho_manh * t["f801c_dinh_luong"] * t["f801c_share_kp"], 4)
        trang_g19 = round(trang_sl * kho_manh * t["taical_dinh_luong"] * taical_share, 4)
        trang_g20 = None

    # ---- Dán stage (glue) — uses Tráng!M9 as the base quantity
    d = CONST["dan"]
    keo_const = d["keo_coeff_opp"] if bao_type == "BOPP" else d["keo_coeff_kp"]
    dan_m9 = trang_sl
    if ov.get("glue") == "9415_60_vistamax_40":
        glue9415 = _exact(["Hạt nhựa PP FC9415"])
        dan_c17 = glue9415["ten"] if glue9415 else "Hạt nhựa PP FC9415"
        dan_d17 = glue9415["ma"] if glue9415 else "NHPPFC9415G01"
        share_a, share_b = d["special_9415_share"], d["special_vistamax_share"]
    else:
        dan_c17 = f801c["ten"] if f801c else "Hạt nhựa nguyên sinh F801C"
        dan_d17 = f801c["ma"] if f801c else "NHPPNSF801C001"
        share_a, share_b = d["f801c_vistamax_share_each"], d["f801c_vistamax_share_each"]
    dan_g17 = round(dan_m9 * share_a * keo_const / 1000, 4)
    dan_g18 = round(dan_m9 * share_b * keo_const / 1000, 4)
    dan_c18 = vistamaxx["ten"] if vistamaxx else "Vistamaxx"
    glue_total_kg = round(dan_g17 + dan_g18, 4)

    # ---- May stage
    chi_may_kg = round(0.6 * qty / 1000, 4)
    day_bo_bao = _exact(["Dây bó bao"])
    day_bo_bao_ten = day_bo_bao["ten"] if day_bo_bao else "Dây bó bao"
    day_bo_bao_kg = round(qty / 5000, 4)

    # ---- Thổi stage (only when the order carries a PE liner bag)
    tui_ldpe_ten = None
    tui_g17 = tui_g18 = None
    tui_ten = tui_ma = None
    quy_cach_day = order.get("tui_long_quy_cach_day") if has_tui_long else None
    day_loai = nvl.quy_cach_day_to_ten(quy_cach_day) if quy_cach_day else None
    if has_tui_long:
        ldpe = _exact(["Hạt nhựa LLDPE", "FD21HN"])
        tui_ldpe_ten = ldpe["ten"] if ldpe else "LDPE"
        if str(order.get("tui_long_loai", "thường")).lower() == "rin":
            tui_g17 = round(qty * ibw * (1 + tol), 4)
            tui_g18 = round(qty * ibw * (1 + tol) * 0.107 * 0, 4)
        else:
            tui_g17 = round(qty * ibw * (1 + tol) * 0.893, 4)
            tui_g18 = round(qty * ibw * (1 + tol) * 0.107, 4)
        rong_cm = order.get("tui_long_rong_cm")
        dai_cm = order.get("tui_long_dai_cm")
        # Tên mô tả theo YCSX — LUÔN dựng sẵn để Thổi!C19 không bao giờ trống.
        # (Template KP có sẵn mã cũ TEB0400521121 ở Thổi!D19 và công thức
        # May!C19 = "=Thổi!C19"; để trống C19 thì sheet May hiện số 0 và sheet
        # Thổi giữ nguyên mã cũ của đơn khác.)
        qc_txt = f", đáy {day_loai}" if day_loai else (f" ({quy_cach_day})" if quy_cach_day else "")
        kt_txt = f"{int(rong_cm)}x{int(dai_cm)}cm " if rong_cm and dai_cm else ""
        tui_ten_ycsx = (f"Túi PE {order.get('tui_long_loai', 'thường')} "
                        f"{kt_txt}{int(round(ibw * 1000))}gr{qc_txt}")
        if order.get("tui_long_ma"):
            tui_ten, tui_ma = order.get("tui_long_ten") or tui_ten_ycsx, order.get("tui_long_ma")
        elif items and rong_cm and dai_cm:
            tui, tui_cands = nvl.find_tui_long(
                items, order.get("tui_long_loai", "thường"), int(rong_cm), int(dai_cm),
                int(round(ibw * 1000)), day_loai=day_loai)
            if tui:
                tui_ten, tui_ma = tui["ten"], tui["ma"]
                if len(tui_cands) > 1:
                    warnings.append(
                        f"Túi PE: Nguyên vật liệu.xlsx có {len(tui_cands)} dòng trùng tên "
                        f"'{tui['ten']}' ({[c['ma'] for c in tui_cands]}) — dùng mã {tui['ma']}, "
                        f"cần kiểm tra lại danh mục NVL.")
                if not day_loai:
                    warnings.append(f"Túi PE: tra được 1 ứng viên ({tui['ten']}, mã {tui['ma']}) "
                                    f"nhưng YCSX không nêu rõ LTMS/MTLS — cần Nhàn xác nhận.")
            else:
                tui_ten, tui_ma = tui_ten_ycsx, None
                warnings.append(f"Chưa xác định được mã Túi PE khớp (rộng/dài/khối lượng/đáy) trong "
                                f"Nguyên vật liệu.xlsx — điền tên theo YCSX ({tui_ten}), để trống Mã, "
                                f"cần Nhàn tra. Ứng viên gần đúng: {[c['ten'] for c in tui_cands]}.")
        elif not items:
            tui_ten, tui_ma = tui_ten_ycsx, None
        else:
            tui_ten, tui_ma = tui_ten_ycsx, None
            warnings.append("YCSX không nêu rõ kích thước túi lồng (rộng/dài) nên chưa tra được Mã "
                            "Túi PE trong Nguyên vật liệu.xlsx — điền tên theo YCSX, để trống Mã, "
                            "cần Nhàn bổ sung.")
    elif bao_type == "BOPP":
        pass  # không lồng túi → ẩn sheet Thổi

    # ---- Nẹp rows (May sheet 21/22)
    # Khối lượng 1 nẹp = rộng nẹp (m) × định lượng nẹp × (rộng bao + 12cm);
    # tổng kg = số lượng × khối lượng 1 nẹp. Một đơn có thể dùng 2 nẹp: nẹp KP
    # để MAY + nẹp giấy để DÁN (YCSX Hùng Duy 25S03IKH00502000046).
    nep_spec = order.get("nep") or _parse_nep(order.get("spec", ""), order)
    nep_extra_m = CONST["may"]["nep_extra_cm"] / 100
    nep_dl = {"KP": CONST["may"]["nep_kp_dinh_luong"],
              "OPP": CONST["may"]["nep_opp_dinh_luong"],
              "PP": CONST["may"]["nep_opp_dinh_luong"],
              "GIAY": CONST["may"]["nep_giay_dinh_luong"]}
    nep_base_kw = {"KP": "nẹp kp", "GIAY": "nẹp giấy", "OPP": "nẹp opp", "PP": "nẹp pp"}
    nep_rows = []
    for nep in nep_spec:
        loai = str(nep.get("loai", "")).upper()
        rong_cm = float(nep.get("rong_cm") or CONST["may"]["nep_rong_cm_default"].get(loai, 6))
        rong_txt = f"{rong_cm:g}".replace(".", ",")
        dinh_luong = nep_dl.get(loai, CONST["may"]["nep_opp_dinh_luong"])
        tokens = [str(t) for t in (nep.get("tokens") or [])
                  if str(t).strip()] or ([str(nep["mau_xuatxu"])] if nep.get("mau_xuatxu") else [])
        name, code = nep.get("ten"), nep.get("ma")
        if not name:
            nep_item, nep_cands = (None, [])
            if items:
                nep_item, nep_cands = nvl.find_nep(
                    items, [nep_base_kw.get(loai, "nẹp")] + tokens, rong_cm)
            if nep_item:
                name, code = nep_item["ten"], nep_item["ma"]
                if len(nep_cands) > 1:
                    warnings.append(
                        f"Nẹp {loai}: Nguyên vật liệu.xlsx có {len(nep_cands)} dòng trùng "
                        f"tên '{nep_item['ten']}' — dùng mã {nep_item['ma']}, cần kiểm tra "
                        f"lại danh mục NVL.")
            else:
                name = f"Nẹp {'giấy' if loai == 'GIAY' else loai} {' '.join(tokens)} {rong_txt}cm".replace("  ", " ")
                code = None
                warnings.append(
                    f"Chưa xác định được mã Nẹp {loai} {' '.join(tokens)} {rong_txt}cm trong "
                    f"Nguyên vật liệu.xlsx — điền tên theo YCSX, để trống Mã, cần Nhàn tra. "
                    f"Ứng viên gần đúng: {[c['ten'] for c in nep_cands] if items else '(chưa có file NVL)'}.")
        if nep.get("rong_cm_default"):
            warnings.append(f"Nẹp {loai}: YCSX không ghi chiều rộng nẹp — dùng mặc định "
                            f"{rong_txt}cm, cần Nhàn xác nhận.")
        if nep.get("tokens_from_bao"):
            warnings.append(f"Nẹp {loai}: YCSX không ghi màu/xuất xứ nẹp — suy theo giấy thân "
                            f"bao ({' '.join(tokens)}), cần Nhàn xác nhận.")
        kg = round((rong_cm / 100) * qty * (W + nep_extra_m) * dinh_luong, 4)
        nep_rows.append({"loai": loai, "name": name, "code": code, "kg": kg})
    if len(nep_rows) > 2:
        warnings.append(f"Đơn dùng {len(nep_rows)} loại nẹp nhưng sheet May chỉ có 2 dòng nẹp "
                        f"(21/22) — chỉ điền 2 dòng đầu, cần Nhàn xử lý tay: "
                        f"{[n['name'] for n in nep_rows[2:]]}.")
    if not nep_rows:
        warnings.append("Đơn không dùng nẹp — sẽ để 0/xóa các ô nẹp ở sheet May.")

    # ---- Tráng 2 (Nhật ký SX)
    trang2 = {
        "kho_tp_mm": int(round(kho_thanh_pham * 1000)),
        "dl_tp": "160 ± 3" if bao_type == "KP" else "110 ± 3",
        "toc_do_may": tsvh.TRANG_TOC_DO_MAY_CO_DINH,
        "dun_keo": tsvh.trang_tra_toc_do_dun_keo(kho_manh_mm),
        "khuon": tsvh.trang_chieu_rong_khuon_xa_keo(kho_manh_mm),
        "f801c_share": t["f801c_share_opp"] if bao_type == "BOPP" else t["f801c_share_kp"],
        "taical_share": taical_share,
        "hat_mau_share": hat_mau_share if bao_type == "BOPP" else None,
    }
    if not trang2["dun_keo"]:
        warnings.append(f"Không tra được 'Tốc độ đùn keo' cho khổ mành {kho_manh_mm}mm — để trống.")

    # ---- Dán 2 (Nhật ký SX)
    length_mm = int(round(L * 1000))
    ngang_mm = int(round(width_m * 1000))
    hong_mm = int(round(gusset_m * 1000))
    dan2_params, tsvh_source = tsvh.dan_cat_tra_thong_so(
        order.get("customer", ""), bao_type, length_mm, ngang_mm, hong_mm)
    if tsvh_source == "chung":
        warnings.append("Dán 2: không có mục TSVH riêng cho khách hàng này — dùng bảng TSVH chung (TSDCA1).")
    missing = [k for k, v in dan2_params.items() if v is None]
    if missing:
        warnings.append(f"Dán 2: các trường {missing} không tra được TSVH — để trống, tô vàng, cần hỏi lại.")
    dan2 = {
        "len_mm": length_mm,
        "d45": f"{length_mm} ± 5",
        "d46": f"{ngang_mm} ± 5",
        "d47": f"{hong_mm} ± 5" if gusset_m > 0 else 0,
        "d37": tsvh.dan_cat_do_sau_dia_xep_hong(hong_mm) if gusset_m > 0 else None,
        "params": dan2_params,
    }
    if gusset_m > 0 and not dan2["d37"]:
        warnings.append(f"Dán 2: không tìm thấy 'Độ sâu dĩa xếp hông' cho hông {hong_mm}mm — để trống.")

    # ---- Thổi 2 (Nhật ký SX)
    thoi2 = None
    if has_tui_long:
        rong_cm = order.get("tui_long_rong_cm")
        dai_cm = order.get("tui_long_dai_cm")
        is_rin = str(order.get("tui_long_loai", "thường")).lower() == "rin"
        thoi2 = {
            "khoi_luong_g": int(round(ibw * 1000)),
            "ldpe_share": 1.0 if is_rin else CONST["thoi"]["ldpe_share"],
            "taical_share": 0.0 if is_rin else CONST["thoi"]["taical_efpe_share"],
        }
        if rong_cm and dai_cm:
            thoi2["rong_mm"] = int(round(rong_cm * 10))
            thoi2["dai_mm"] = int(round(dai_cm * 10))
            thoi2["do_day"] = round(ibw * 1000 * 10000 / rong_cm / dai_cm / 2 / 0.93)
            if day_loai == "dài" or order.get("tui_long_may_dinh_day"):
                thoi2["day"] = "40-50"
            elif day_loai == "ngắn":
                thoi2["day"] = "20-30"
            else:
                thoi2["day"] = None
                warnings.append("Kích thước đáy (Thổi 2): quy cách túi lồng không rõ LTMS/MTLS — để trống.")
            thoi2["dun_keo"], thoi2["keo_bong"] = tsvh.thoi_tra_toc_do(thoi2["rong_mm"])
            if not thoi2["dun_keo"]:
                warnings.append(f"Không tìm thấy khoảng khổ {thoi2['rong_mm']}mm trong TSVH THỔI — để trống.")
        else:
            warnings.append("Thổi 2: thiếu rộng/dài túi lồng nên chưa điền kích thước/độ dày — để trống.")

    # ---- Size text for In 1 (Kích thước SX / TP), format "320+80 x 680" per YCSX
    def _kt(w_g, g_m, l_m):
        ngang = int(round((w_g - g_m) * 1000))
        hong = int(round(g_m * 1000))
        dai = int(round(l_m * 1000))
        return f"{ngang}+{hong} x {dai}" if hong else f"{ngang} x {dai}"

    size_sx_mm = _kt(W, gusset_m, L)
    w_tp = float(order.get("width_plus_gusset_tp_m", W))
    g_tp = float(order.get("gusset_tp_m", gusset_m))
    l_tp = float(order.get("length_tp_m", L))
    size_tp_mm = _kt(w_tp, g_tp, l_tp)

    # ---- Đóng gói: số bao/kiện theo loại bao + chiều dài (bảng TSVH ĐÓNG GÓI)
    bao_kien = _bao_kien(bao_type, int(round(L * 100)))
    bao_kien_text = f"{bao_kien} bao/kiện" if bao_kien else None

    fields = {
        "customer": order.get("customer", ""), "product_name": order.get("product_name", ""),
        "product_code": order.get("product_code", ""), "order_id": order.get("order_id", ""),
        "so_phieu_sx": order.get("so_phieu_sx", "SX" + str(order.get("order_id", ""))[-6:]),
        "qty": int(qty) if qty == int(qty) else qty,
        "bag_length_m": L, "width_plus_gusset_m": W,
        "tolerance": tol, "bao_type": bao_type,
        "has_print": has_print, "has_tui_long": has_tui_long,
        "so_mau_in": so_mau_in if has_print else 0,
        "inner_bag_weight_kg": ibw if has_tui_long else None,
        "gusset_m": gusset_m, "width_m": width_m,
        "kho_manh": kho_manh, "kho_mang": kho_mang if bao_type == "BOPP" else None,
        "kho_giay": kho_giay if bao_type == "KP" else None,
        "sl_in_thuc_te_m": sl_in, "trang_sl": trang_sl,
        "thanh_pham_du_kien": thanh_pham,
        "bag_type_label": "Bao OPP" if family == "opp" else "Bao KP",
        "size_sx_mm": size_sx_mm, "size_tp_mm": size_tp_mm,
        "sl_theo_po_m": trang_sl,
        # NVL names/codes
        "manh_ten": manh["ten"] if manh else None, "manh_ma": manh["ma"] if manh else None,
        "manh_kg": trang_g17,
        "f801c_ten": f801c["ten"] if f801c else "F801C",
        "taical_ten": taical["ten"] if taical else "Taical",
        "trang_g18": trang_g18, "trang_g19": trang_g19, "trang_g20": trang_g20,
        "mang_ten": mang["ten"] if mang else None, "mang_ma": mang["ma"] if mang else None,
        "giay_ten": giay["ten"] if giay else None, "giay_ma": giay["ma"] if giay else None,
        "mang_bopp_kg": mang_bopp_kg,
        "dung_moai_opp_kg": dung_moai_opp_kg, "dung_moai_ea_kg": dung_moai_ea_kg,
        "giay_kraft_kg": giay_kraft_kg,
        "giay_kraft_code": giay["ma"] if giay else order.get("giay_kraft_code", "GN07010200001"),
        "kraft_name": giay["ten"] if giay else (order.get("kraft_name")
                     or f"Giấy Kraft {order.get('giay_kraft_mau','vàng')} "
                        f"{order.get('giay_kraft_xuatxu','Nhật')} "
                        f"K{int(round(kho_giay*1000))} ĐL70"),
        # Dán
        "dan_c17": dan_c17, "dan_d17": dan_d17, "dan_c18": dan_c18,
        "dan_g17": dan_g17, "dan_g18": dan_g18, "glue_total_kg": glue_total_kg,
        # May
        "chi_may_kg": chi_may_kg, "day_bo_bao_ten": day_bo_bao_ten, "day_bo_bao_kg": day_bo_bao_kg,
        "nep_rows": nep_rows,
        # Thổi
        "tui_ldpe_ten": tui_ldpe_ten, "tui_g17": tui_g17, "tui_g18": tui_g18,
        "tui_ten": tui_ten, "tui_ma": tui_ma,
        "quy_cach_day": quy_cach_day, "day_loai": day_loai,
        # Đóng gói
        "bao_kien": bao_kien, "bao_kien_text": bao_kien_text,
        # Tráng 2 / Dán 2 / Thổi 2
        "trang2": trang2, "dan2": dan2, "thoi2": thoi2,
        "warnings": warnings,
    }
    return fields


def _dims_lookup(order, key):
    w = order.get("width_cm"); g = order.get("gusset_cm")
    if not w or not g:
        return 0
    for row in DIMS["rows"]:
        if row["width_cm"] == int(w) and row["gusset_cm"] == int(g):
            return row[key] / 100
    return 0


# -------------------------------------------------------------------- fillers
def _set(xp, sheet, ref, value):
    if value is None or value == "":
        return
    xp.set_value(sheet, ref, value)


def _blank(xp, sheet, ref):
    xp.set_value(sheet, ref, None)


def _clear_ink_rows(xp, start, end):
    """Mực in được công đoạn In điền thủ công theo thiết kế — xóa hết dòng mực
    (tên/mã/STT/ĐVT/số lượng) để sheet In không còn tên mực leftover từ template."""
    for r in range(start, end + 1):
        for col in "BCDFG":
            _blank(xp, "In", f"{col}{r}")


def fill_dinh_muc(template_path, family, fields, so_mau_in, out_path):
    xp = XlsxPatch(template_path)
    cm = CELLMAP["dinh_muc"]
    bao_type = fields["bao_type"]
    has_print = fields["has_print"]
    has_tui_long = fields["has_tui_long"]

    def present(sheet):
        return sheet in xp.sheet_paths

    # --- In sheet: header + info tables ---
    for region in ("header", "info_table_M", "info_table_C"):
        for ref, field in cm["In"].get(region, {}).items():
            _set(xp, "In", ref, fields.get(field))
    _set(xp, "In", "G29", fields.get("thanh_pham_du_kien"))

    # --- In sheet: materials (print) ---
    if has_print:
        if bao_type == "BOPP":
            _set(xp, "In", "C18", fields.get("mang_ten"))
            _set(xp, "In", "D18", fields.get("mang_ma"))
            _set(xp, "In", "F18", "Kg")
            _set(xp, "In", "G18", fields.get("mang_bopp_kg"))
            _set(xp, "In", "G19", fields.get("sl_in_thuc_te_m"))
            _set(xp, "In", "G20", fields.get("dung_moai_opp_kg"))
            _set(xp, "In", "G21", fields.get("dung_moai_ea_kg"))
            _clear_ink_rows(xp, 22, 28)
        else:
            _set(xp, "In", "C18", fields.get("giay_ten") or fields.get("kraft_name"))
            _set(xp, "In", "D18", fields.get("giay_kraft_code"))
            _set(xp, "In", "F18", "Kg")
            _set(xp, "In", "G18", fields.get("giay_kraft_kg"))
            _set(xp, "In", "G19", fields.get("sl_in_thuc_te_m"))
            _clear_ink_rows(xp, 20, 23)
    else:
        # không in: blank leftover G formulas so the (hidden) sheet carries no stale data
        if bao_type == "BOPP":
            for ref in ("G18", "G19", "G20", "G21"):
                xp.set_value("In", ref, None)
            _clear_ink_rows(xp, 22, 28)
        else:
            _clear_ink_rows(xp, 20, 23)

    # --- In 2 (materials mirror) ---
    if has_print and present("In 2"):
        _set(xp, "In 2", "B18", fields.get("mang_ten") or fields.get("giay_ten") or fields.get("kraft_name"))
        _set(xp, "In 2", "C18", fields.get("mang_ma") or fields.get("giay_kraft_code"))
        if bao_type == "BOPP":
            _set(xp, "In 2", "B20", "Dung môi OPP")
            _set(xp, "In 2", "C20", "PDMTOLUENE001")
            _set(xp, "In 2", "B21", "Dung môi EA")
            _set(xp, "In 2", "C21", "PDMEA00001")

    # --- Stage material sheets: header + info table M ---
    for sheet in ("Tráng", "Dán", "Thổi", "May"):
        if not present(sheet):
            continue
        for region in ("header", "info_table_M"):
            for ref, field in cm.get(sheet, {}).get(region, {}).items():
                _set(xp, sheet, ref, fields.get(field))

    # --- Tráng ---
    _set(xp, "Tráng", "M9", fields.get("trang_sl"))
    _set(xp, "Tráng", "M10", fields.get("kho_manh"))
    if bao_type == "BOPP":
        _set(xp, "Tráng", "M11", fields.get("kho_mang"))
    else:
        _set(xp, "Tráng", "M12", fields.get("kho_giay"))
    _set(xp, "Tráng", "C17", fields.get("manh_ten"))
    _set(xp, "Tráng", "D17", fields.get("manh_ma"))
    _set(xp, "Tráng", "G17", fields.get("manh_kg"))
    _set(xp, "Tráng", "C18", fields.get("f801c_ten"))
    _set(xp, "Tráng", "C19", fields.get("taical_ten"))
    _set(xp, "Tráng", "G18", fields.get("trang_g18"))
    _set(xp, "Tráng", "G19", fields.get("trang_g19"))
    if bao_type == "BOPP":
        _set(xp, "Tráng", "G20", fields.get("trang_g20"))
    if not has_print:
        # vật liệu chính chuyển xuống Tráng (dòng STT4)
        if bao_type == "KP":
            _set(xp, "Tráng", "C20", fields.get("giay_ten") or fields.get("kraft_name"))
            _set(xp, "Tráng", "D20", fields.get("giay_kraft_code"))
            _set(xp, "Tráng", "F20", "Kg")
            _set(xp, "Tráng", "G20", fields.get("giay_kraft_kg"))
        else:
            _set(xp, "Tráng", "C20", fields.get("mang_ten"))
            _set(xp, "Tráng", "D20", fields.get("mang_ma"))
            _set(xp, "Tráng", "F20", "Kg")
            _set(xp, "Tráng", "G20",
                 round(fields["trang_sl"] * (fields["kho_mang"] or 0) * 0.01584, 4))

    # --- Dán ---
    _set(xp, "Dán", "M10", fields.get("kho_manh"))
    _set(xp, "Dán", "C17", fields.get("dan_c17"))
    _set(xp, "Dán", "D17", fields.get("dan_d17"))
    _set(xp, "Dán", "G17", fields.get("dan_g17"))
    _set(xp, "Dán", "C18", fields.get("dan_c18"))
    _set(xp, "Dán", "G18", fields.get("dan_g18"))

    # --- Thổi (túi lồng) ---
    # C19/D19 ghi thẳng qua set_value (KHÔNG qua _set) để None cũng được ghi:
    # template KP mang sẵn mã túi cũ TEB0400521121 ở D19, _set() bỏ qua None nên
    # mã của đơn khác sẽ lọt vào phiếu.
    if has_tui_long and present("Thổi"):
        _set(xp, "Thổi", "C17", fields.get("tui_ldpe_ten"))
        _set(xp, "Thổi", "G17", fields.get("tui_g17"))
        _set(xp, "Thổi", "G18", fields.get("tui_g18"))
        xp.set_value("Thổi", "C19", fields.get("tui_ten") or None)
        xp.set_value("Thổi", "D19", fields.get("tui_ma") or None)
    elif present("Thổi"):
        # không lồng túi: sheet Thổi bị ẩn, nhưng May!C19/D19/G19/G20 là công thức
        # "=Thổi!…" nên vẫn hiện 0 + mã cũ trên sheet May đang hiển thị → xóa cả hai.
        for ref in ("C19", "D19"):
            xp.set_value("Thổi", ref, None)
        for ref in ("C19", "D19", "G19", "G20"):
            xp.set_value("May", ref, None)

    # --- May ---
    _set(xp, "May", "G17", fields.get("chi_may_kg"))
    _set(xp, "May", "C18", fields.get("day_bo_bao_ten"))
    _set(xp, "May", "G18", fields.get("day_bo_bao_kg"))
    nep_rows = fields.get("nep_rows", [])
    for i, r in enumerate((21, 22)):
        nep = nep_rows[i] if i < len(nep_rows) else None
        if nep:
            xp.set_value("May", f"C{r}", nep["name"])
            xp.set_value("May", f"D{r}", nep["code"] or None)
            _set(xp, "May", f"F{r}", "Kg")   # template OPP thiếu ĐVT ở dòng 21/22
            _set(xp, "May", f"G{r}", nep["kg"])
        else:
            _blank(xp, "May", f"C{r}")
            _blank(xp, "May", f"D{r}")
            _set(xp, "May", f"G{r}", 0)

    # --- May 2 (Quy cách) ---
    if present("May 2"):
        cell_qc = "C37" if bao_type == "BOPP" else "C35"
        if has_tui_long:
            _set(xp, "May 2", cell_qc, fields.get("quy_cach_day"))
        else:
            _blank(xp, "May 2", cell_qc)

    # --- In 1 — kích thước SX/TP (Lệnh SX sheets pull from here by formula) ---
    if present("In 1"):
        _set(xp, "In 1", "C17", fields.get("size_sx_mm"))
        _set(xp, "In 1", "C18", fields.get("size_tp_mm"))

    # --- Tráng 2 ---
    if present("Tráng 2"):
        if bao_type == "KP":
            _set(xp, "Tráng 2", "B21", fields.get("manh_ten"))
            _set(xp, "Tráng 2", "C21", fields.get("manh_ma"))
            if not has_print:
                _set(xp, "Tráng 2", "A24", 5)
                _set(xp, "Tráng 2", "B24", fields.get("giay_ten") or fields.get("kraft_name"))
                _set(xp, "Tráng 2", "C24", fields.get("giay_kraft_code"))
                _set(xp, "Tráng 2", "F24", "Kg")
                _set(xp, "Tráng 2", "G24", fields.get("giay_kraft_kg"))
        t2 = fields["trang2"]
        cells2 = {"KP":   {"kho_tp": "M32", "dl_tp": "M33", "toc_do": "D31", "dun_keo": "D50", "khuon": "D52"},
                  "BOPP": {"kho_tp": "M31", "dl_tp": "M32", "toc_do": "D30", "dun_keo": "D49", "khuon": "D51"}}[bao_type]
        _set(xp, "Tráng 2", cells2["kho_tp"], t2["kho_tp_mm"])
        _set(xp, "Tráng 2", cells2["dl_tp"], t2["dl_tp"])
        _set(xp, "Tráng 2", cells2["toc_do"], t2["toc_do_may"])
        _set(xp, "Tráng 2", cells2["dun_keo"], t2["dun_keo"])
        _set(xp, "Tráng 2", cells2["khuon"], t2["khuon"])
        # Tỷ lệ vật liệu (cột E): OPP E21=F801C/E22=Taical/E23=Hạt màu; KP E22=F801C/E23=Taical
        if bao_type == "BOPP":
            _set(xp, "Tráng 2", "E21", t2["f801c_share"])
            _set(xp, "Tráng 2", "E22", t2["taical_share"])
            _set(xp, "Tráng 2", "E23", t2["hat_mau_share"])
        else:
            _set(xp, "Tráng 2", "E22", t2["f801c_share"])
            _set(xp, "Tráng 2", "E23", t2["taical_share"])

    # --- Dán 2 ---
    if present("Dán 2"):
        dan2 = fields["dan2"]
        _set(xp, "Dán 2", "D34", dan2["len_mm"])
        _set(xp, "Dán 2", "D45", dan2["d45"])
        _set(xp, "Dán 2", "D46", dan2["d46"])
        _set(xp, "Dán 2", "D47", dan2["d47"])
        _set(xp, "Dán 2", "D30", dan2["params"].get("toc_do_may"))
        _set(xp, "Dán 2", "D31", dan2["params"].get("toc_do_keo"))
        _set(xp, "Dán 2", "D32", dan2["params"].get("lai_cuon"))
        _set(xp, "Dán 2", "D33", dan2["params"].get("luc_ep_duong_dan"))
        _set(xp, "Dán 2", "D36", dan2["params"].get("dieu_chinh_duong"))
        if dan2["d37"] is not None:
            _set(xp, "Dán 2", "D37", dan2["d37"])
        else:
            _blank(xp, "Dán 2", "D37")

    # --- Thổi 2 ---
    if has_tui_long and present("Thổi 2") and fields.get("thoi2"):
        t2 = fields["thoi2"]
        cells = {"KP":   {"rong": "D43", "dai": "D44", "day": "D45", "do_day": "D46",
                          "khoi_luong": "D47", "dun_keo": "D36", "keo_bong": "D37"},
                 "BOPP": {"rong": "D45", "dai": "D46", "day": "D47", "do_day": "D48",
                          "khoi_luong": "D49", "dun_keo": "D39", "keo_bong": "D40"}}[bao_type]
        if "rong_mm" in t2:
            _set(xp, "Thổi 2", cells["rong"], f"{t2['rong_mm']} ± 10")
            _set(xp, "Thổi 2", cells["dai"], f"{t2['dai_mm']} ± 10")
            _set(xp, "Thổi 2", cells["day"], t2.get("day"))
            _set(xp, "Thổi 2", cells["do_day"], f"{t2['do_day']} ± 1")
        _set(xp, "Thổi 2", cells["khoi_luong"], f"{t2['khoi_luong_g']} ± 1")
        _set(xp, "Thổi 2", cells["dun_keo"], t2.get("dun_keo"))
        _set(xp, "Thổi 2", cells["keo_bong"], t2.get("keo_bong"))
        # Tỷ lệ vật liệu (cột E): LDPE / Taical EFPE (rin = 100% LDPE, không taical)
        _set(xp, "Thổi 2", "E18", t2.get("ldpe_share"))
        _set(xp, "Thổi 2", "E19", t2.get("taical_share"))

    # --- Đóng gói 1 & 2 (Lệnh SX + Nhật ký SX) ---
    for sn in ("Đóng gói 1", "Đóng gói 2"):
        if not present(sn):
            continue
        for region in ("header", "stage"):
            for ref, field in cm.get(sn, {}).get(region, {}).items():
                if ref.startswith("_"):
                    continue
                _set(xp, sn, ref, fields.get(field))
        if sn == "Đóng gói 2" and not fields.get("bao_kien_text"):
            _blank(xp, sn, "C32")
        # xóa dữ liệu đơn hàng cũ sót lại trong template
        for ref in ("V5", "W5", "X5"):
            _blank(xp, sn, ref)

    # --- Sheet X + hidden sheets ---
    if not has_print and present("X"):
        xp.clear_sheet_images("X")
    if not has_print:
        for sn in ("In", "In 1", "In 2"):
            if present(sn):
                xp.set_sheet_state(sn, "hidden")
    if not has_tui_long:
        for sn in ("Thổi", "Thổi 1", "Thổi 2"):
            if present(sn):
                xp.set_sheet_state(sn, "hidden")

    xp.save(out_path)
    return out_path


def _single_product(order):
    """Collapse a flat (single-product) order dict into one product record."""
    return {
        "product_code": order.get("product_code", ""),
        "product_name": order.get("product_name", ""),
        "ma_code": order.get("ma_code", ""),
        "qty": order.get("qty"),
        "spec": order.get("spec", ""),
    }


def _per_product_order(order, product):
    """Build a compute()-ready order for one product (inherits header + shared dims)."""
    po = {k: v for k, v in order.items() if k != "products"}
    po.update(product)
    po.update(parse_spec(product.get("spec", "")))
    if not po.get("inner_bag_weight_kg"):
        ibw = _extract_pe_liner_weight(product.get("product_name", ""))
        if ibw is not None:
            po["inner_bag_weight_kg"] = ibw
    return po


def fill_ycsx(src_path, order, out_path, preserve=False):
    """Fill the YCSX form: header (label + value) + ALL product line items.

    preserve=True — the input IS a YCSX file (clone-and-fill): the source
    workbook stays intact apart from the header block and the parsed product
    lines, so mã code, mã khách hàng, giao-hàng notes and the per-stage quality
    block (MÀNH/IN/TRÁNG/DÁN/THỔI/MAY) are never deleted. Hai quy tắc giữ phiếu
    của sale nguyên vẹn (sửa 2026-08-12):
      * chỉ ghi ô nào giá trị THỰC SỰ khác so với file gốc (_write/_same_cell);
      * ghi vào ĐÚNG toạ độ ô đã dò được khi parse (order["ycsx_header_cells"]),
        không ghi cứng B6:B10 — layout của mỗi khách một khác, ghi cứng sẽ tạo
        thêm một dòng "5. Số đơn hàng" thứ hai ngay trên bảng sản phẩm.
    preserve=False — no input YCSX (order JSON / sample, generic template):
    the stale product region is cleared so leftovers can't leak into the output.

    Stale template data (leftover products, stage markers, old header values,
    old giao-hàng/notes) is cleared so output never leaks a previous order. The
    shared spec is written into each product row's F cell; where F is the
    top-left of a merge (e.g. F13:H15) that sets the merged display.

    Note: The template has a single merge F13:H15 covering rows 13-15.
    For multi-product orders, products in rows 14+ cannot have their own
    spec cell (it's inside the merge). set_value returns False for those
    cells — the spec is still fully detailed in each product's Định mức file.
    """
    xp = XlsxPatch(src_path)
    cm = CELLMAP["ycsx"]
    sheet = list(xp.sheet_paths)[0]
    snapshot = order.get("ycsx_snapshot") or {}

    def _write(ref, value):
        """Ghi ô — nhưng ở chế độ preserve thì BỎ QUA khi giá trị không đổi.

        Đơn Hùng Duy 25S03IKH00502000046: mọi giá trị đều vừa được đọc ra từ chính
        file này, nên ghi lại là vô nghĩa mà vẫn có rủi ro làm phiếu của sale khác đi.
        Không ghi gì thì không thể nào tạo thêm dòng trùng.
        """
        if preserve and ref in snapshot and _same_cell(snapshot[ref], value):
            return False
        return xp.set_value(sheet, ref, value)

    # header — ghi vào ĐÚNG ô đã dò được khi parse (không ghi cứng B6:B10)
    header_cells = order.get("ycsx_header_cells") or {}
    if header_cells:
        labels = {v["ref"]: v["label"] for v in header_cells.values()}
        hfields = {v["ref"]: k for k, v in header_cells.items()}
    else:  # nguồn là JSON/sample/template chung — dùng layout mặc định QTKSX-BM01
        labels = cm.get("header_labels") or {
            "B6": "1. Ngày yêu cầu", "B7": "2. Khách hàng", "B8": "3. Địa chỉ",
            "B9": "4. Mã khách hàng", "B10": "5. Số đơn hàng",
        }
        hfields = {"B6": "ngay_yc", "B7": "customer", "B8": "address",
                   "B9": "customer_code", "B10": "order_id"}
    for ref, label in labels.items():
        val = order.get(hfields.get(ref, ""), "") or ""
        _write(ref, f"{label}: {val}".rstrip())

    # xóa các ô TRÙNG nhãn header nằm phía trên bảng sản phẩm (bản sao thừa của
    # "5. Số đơn hàng: …" mà phiếu không cần)
    if preserve:
        for ref in order.get("ycsx_dup_header_cells") or []:
            xp.set_value(sheet, ref, None)

    # product line items below the (parsed) header row
    products = order.get("products") or [_single_product(order)]
    ycsx_cols = order.get("ycsx_cols", {})
    base_row = int(order.get("ycsx_header_row") or 12)
    start_row = base_row + 1
    col_code = ycsx_cols.get("code", "C")
    col_name = ycsx_cols.get("name", "D")
    col_ma = ycsx_cols.get("ma", "E")
    col_spec = ycsx_cols.get("spec", "F")
    col_qty = ycsx_cols.get("qty", "J")
    col_dvt = ycsx_cols.get("dvt") or (None if preserve else "I")

    last_data_row = base_row
    for i, p in enumerate(products):
        r = start_row + i
        _write(f"B{r}", i + 1)
        _write(f"{col_code}{r}", p.get("product_code", ""))
        _write(f"{col_name}{r}", p.get("product_name", ""))
        # ma_code/spec: ở chế độ preserve chúng vừa được đọc ra từ chính file này
        # (và spec hay nằm trong merge F13:H15) → không ghi lại.
        if not preserve:
            xp.set_value(sheet, f"{col_ma}{r}", p.get("ma_code", ""))
            xp.set_value(sheet, f"{col_spec}{r}", p.get("spec", ""))
        if col_dvt:
            _write(f"{col_dvt}{r}", "Cái")
        _write(f"{col_qty}{r}", p.get("qty", ""))
        last_data_row = r

    if not preserve:
        # generic template: clear stale notes/schedule (K,L) across the whole
        # product region, then clear B-J beyond the products written (leftover
        # products + stage markers at 16-21 + giao-hàng at 23). Merged
        # top-lefts (e.g. B23) clear on contact. Cells inside a merge range
        # don't exist as individual <c> elements — set_value returns False for
        # those, which is harmless (no stale data can live in a non-existent cell).
        for r in range(start_row, 24):
            xp.set_value(sheet, f"K{r}", "")
            xp.set_value(sheet, f"L{r}", "")
        for r in range(last_data_row + 1, 24):
            for col in "BCDEFGHIJ":
                xp.set_value(sheet, f"{col}{r}", "")

    xp.save(out_path)
    return out_path


# --------------------------------------------------------------------- sample
SAMPLE_40KG = {
    "customer": "CÔNG TY 4 ORANGES CO.,LTD",
    "product_name": "4O BTA140 SPEC CEO PUTTY FOR INTERIOR 40KG",
    "product_code": "XD1KH00565BP0011",
    "order_id": "26BA2HKH00565000014",
    "so_phieu_sx": "SX2607080",
    "qty": 3000, "bag_length_m": 0.82, "width_plus_gusset_m": 0.5,
    "width_cm": 42, "gusset_cm": 8, "inner_bag_weight_kg": 0.051,
    "kho_mang": 1.04, "spec": "1.Kích thước: (42+8) cm x 82cm\n3. Cấu trúc: Bao BOPP in ống đồng - OPP mờ",
}


def run(source, colors, outdir=None):
    """Generate one Định mức per product + one shared YCSX.

    source: ("sample", None) | ("order", path) | ("ycsx", path) | ("dict", order_dict)
    Returns {"family", "outdir", "outputs":[paths], "products":[{product_name, fields}]}.
    Reused by the CLI (main) and by the web/claude.ai code-execution wrapper.
    """
    kind = source[0]
    if kind == "sample":
        order = copy.deepcopy(SAMPLE_40KG)
    elif kind == "order":
        with open(source[1], encoding="utf-8") as f:
            order = json.load(f)
    elif kind == "ycsx":
        order = parse_ycsx(source[1])
    elif kind == "dict":
        order = copy.deepcopy(source[1])
    else:
        raise ValueError(f"bad source kind {kind!r}")

    products = order.get("products") or [_single_product(order)]
    order["products"] = products
    family = order.get("bag_family") or detect_family(order)

    outdir = outdir or os.path.join(OUTDIR, datetime.date.today().isoformat())
    os.makedirs(outdir, exist_ok=True)
    tmpl = template_path(family)

    outputs, summary = [], []
    for product in products:
        porder = _per_product_order(order, product)
        validate_inputs(porder, family, colors)
        fields = compute(porder, family, colors)
        safe = re.sub(r"[^0-9A-Za-z]+", "-", product.get("product_name") or "order")[:40]
        dm_out = os.path.join(outdir, f"Định mức - {safe} - {order.get('order_id','')}.xlsx")
        fill_dinh_muc(tmpl, family, fields, colors, dm_out)
        outputs.append(dm_out)
        summary.append({"product_name": product.get("product_name"), "fields": fields})

    ycsx_safe = re.sub(r"[^0-9A-Za-z]+", "-",
                       order.get("product_name") or products[0].get("product_name") or "order")[:40]
    # YCSX: khi input là chính file YCSX thì clone file gốc (giữ mã code/mã
    # khách hàng/chất lượng yêu cầu các công đoạn); JSON/sample dùng template.
    ycsx_out = os.path.join(outdir, f"YCSX - {ycsx_safe} - {order.get('order_id','')}.xlsx")
    ycsx_base = source[1] if kind == "ycsx" else template_path("ycsx")
    fill_ycsx(ycsx_base, order, ycsx_out, preserve=kind == "ycsx")
    outputs.append(ycsx_out)

    zip_path = os.path.join(outdir, f"Dinh_muc_YCSX_{order.get('order_id','order')}.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fpath in outputs:
            zf.write(fpath, os.path.basename(fpath))

    return {"family": family, "outdir": outdir, "outputs": [zip_path], "products": summary}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--order", help="path to order JSON")
    ap.add_argument("--ycsx", help="path to a YCSX .xlsx to parse")
    ap.add_argument("--sample", action="store_true", help="use built-in 40KG sample")
    ap.add_argument("--colors", type=int, help="pre-answer số màu in (skips the prompt)")
    ap.add_argument("--outdir", help="output directory (default: output/<today/>)")
    args = ap.parse_args()

    if args.sample:
        source = ("sample", None)
    elif args.order:
        source = ("order", args.order)
    elif args.ycsx:
        source = ("ycsx", args.ycsx)
    else:
        ap.error("provide --order, --ycsx, or --sample")

    if args.colors is not None:
        so_mau_in = args.colors
    else:
        so_mau_in = 0
        while so_mau_in < 0:
            raw = input("Số màu in? (number of print colors — nhập 0 nếu không in): ").strip()
            if raw:
                try:
                    so_mau_in = int(raw)
                except ValueError:
                    pass
            if so_mau_in < 0:
                print("Số màu in không được âm. Vui lòng nhập lại.", file=sys.stderr)
    if so_mau_in < 0:
        print("Số màu in không được âm.", file=sys.stderr)
        sys.exit(2)

    try:
        res = run(source, so_mau_in, outdir=args.outdir)
    except InputValidationError as e:
        print(e, file=sys.stderr)
        sys.exit(2)
    print(f"Detected bag family: {res['family']}  (template: {BAG['families'][res['family']]['_aka']})")
    print(f"số màu in = {so_mau_in}\n")
    for p in res["products"]:
        f = p["fields"]
        print(f"== {p['product_name']} ==")
        for k in ("qty", "bag_length_m", "width_plus_gusset_m", "sl_in_thuc_te_m", "so_mau_in",
                  "kho_manh", "kho_mang", "kho_giay", "inner_bag_weight_kg"):
            print(f"  {k:24s} = {f.get(k)}")
        for k in ("mang_bopp_kg", "dung_moai_opp_kg", "dung_moai_ea_kg", "giay_kraft_kg", "glue_total_kg"):
            if k in f:
                print(f"  {k:24s} = {f.get(k)}")
    print("\nGenerated files:")
    for path in res["outputs"]:
        print("  ✓", path)


if __name__ == "__main__":
    main()

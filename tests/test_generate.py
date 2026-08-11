"""Tests: every expected cell filled, regression values, hidden-sheet logic."""
from __future__ import annotations

import json
import os
import sys
import tempfile

from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from generate import compute, detect_family, run

HERE = os.path.dirname(__file__)
SAMPLES = os.path.join(HERE, "..", "samples")


def _load_json(name):
    with open(os.path.join(SAMPLES, name), encoding="utf-8") as f:
        return json.load(f)


SAMPLE_25KG = _load_json("25kg_tan_chau.json")
SAMPLE_40KG = _load_json("40kg_4oranges.json")

SAMPLE_40KG_NORMAL = {
    "_comment": "Normal (non-4O) OPP order with a PE-rin liner. Expected: Tráng taical 10% / hạt màu 5% "
                "(not 7/8 — that's only 4 Oranges/Thanh Phụng); Thổi rin = 100% LDPE, 0 taical; "
                "Đóng gói OPP dài 82cm = 300 bao/kiện.",
    "customer": "CÔNG TY CỔ PHẦN HÃNG SƠN ĐÔNG Á",
    "product_name": "Bao BT Waler Nội Thất Cao Cấp 40kg",
    "product_code": "XD1KH00069BP0012",
    "order_id": "26BA2IKH00010000004",
    "so_phieu_sx": "SX000004",
    "qty": 5000, "bag_length_m": 0.82, "width_plus_gusset_m": 0.5,
    "width_cm": 42, "gusset_cm": 8, "inner_bag_weight_kg": 0.02, "kho_mang": 1.04,
    "spec": ("1.Kích thước: (42+8) x 82 cm (±1)\n2.Thành phẩm: (42+8) x 79cm (±1)\n"
             "3. Cấu trúc: Bao BOPP in ống đồng - Màng BOPP bóng - Ghép PP dệt trắng\n"
             "5. Quy cách lồng túi PE: Lồng túi PE rin 50x92 (20gr) (LTMS)\n"
             "6. Quy cách đóng gói: Đóng gói theo tiêu chuẩn BBAS."),
}


def _num(ws, cell):
    v = ws[cell].value
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _str(ws, cell):
    v = ws[cell].value
    return str(v).strip() if v is not None else ""


def _check_all_filled(ws, cells, label):
    """Assert every cell in the list is non-empty."""
    for cell in cells:
        v = _str(ws, cell)
        assert v != "", f"{label}: {cell} should be filled but is empty"


def _check_no_solid_fill(ws, cells, label):
    """Assert no cell has a solid fill colour (cleanliness guard)."""
    BAD = ("FFFF00", "FF0000")
    for cell in cells:
        c = ws[cell]
        if c.fill and c.fill.fgColor:
            rgb = c.fill.fgColor.rgb
            if rgb and rgb not in ("00000000",):
                assert rgb not in BAD, f"{label}: {cell} has forbidden fill {rgb}"


# ---- sheet: In ----
IN_HEADER = [f"C{i}" for i in range(5, 10)]
IN_INFO_M_KP = [f"M{i}" for i in range(6, 12)] + ["M13", "M14"]
IN_INFO_M_OPP = [f"M{i}" for i in range(6, 13)] + ["M14"]
IN_INFO_C = [f"C{i}" for i in range(39, 48)]
IN_THANH_PHAM = ["G29"]


def _make_fixture(order, colors, prefix):
    outdir = tempfile.mkdtemp(prefix=prefix)
    family = detect_family(order)
    fields = compute(order, family, colors)
    run(("dict", order), colors, outdir=outdir)
    xlsx_files = [os.path.join(outdir, f) for f in os.listdir(outdir)
                  if f.endswith(".xlsx") and f.startswith("Định mức")]
    assert xlsx_files, f"no Dinh muc xlsx in {os.listdir(outdir)}"
    wb = load_workbook(xlsx_files[0], data_only=True)
    return family, fields, wb


def _sheet_state(wb, name):
    return wb[name].sheet_state if name in wb.sheetnames else None


class Test25KgKp:
    """Paper/KP 25kg Tan Chau -- 2 mau in."""

    @classmethod
    def setup_class(cls):
        cls.family, cls.fields, cls.wb = _make_fixture(dict(SAMPLE_25KG), 2, "dm_test_25kg_")
        cls.ws = cls.wb["In"]

    @classmethod
    def teardown_class(cls):
        cls.wb.close()

    def test_completeness_header(self):
        _check_all_filled(self.ws, IN_HEADER, "In header")

    def test_completeness_info_m(self):
        _check_all_filled(self.ws, IN_INFO_M_KP, "In info_M")

    def test_completeness_thanh_pham(self):
        _check_all_filled(self.ws, IN_THANH_PHAM, "In thanh_pham")

    def test_kraft_material_filled(self):
        _check_all_filled(self.ws, ["C18", "D18", "G18"], "In kraft")

    def test_kraft_name_filled(self):
        assert "K1020" in _str(self.ws, "C18"), f"In!C18 got {_str(self.ws, 'C18')!r}"
        assert _str(self.ws, "D18") == "GN07010200001", f"D18={_str(self.ws, 'D18')!r}"

    def test_stage_values_filled(self):
        assert _num(self.ws, "M9") == 4600.0, f"M9={_num(self.ws, 'M9')}"
        for sheet, cells in {
            "Tráng": ["C17", "D17", "G17", "C18", "C19", "G18", "G19"],
            "Dán":   ["C17", "D17", "G17", "C18", "G18"],
            "May":   ["G17", "G18", "C18"],
        }.items():
            _check_all_filled(self.wb[sheet], cells, f"{sheet} stage")

    def test_dan_glue(self):
        ws = self.wb["Dán"]
        assert abs(_num(ws, "G17") - 9.1466) < 0.01, f"Dán!G17={_num(ws, 'G17')}"
        assert abs(_num(ws, "G18") - 6.0978) < 0.01, f"Dán!G18={_num(ws, 'G18')}"

    def test_trang_manh(self):
        ws = self.wb["Tráng"]
        assert _str(ws, "D17") == "MW07010600001", f"Tráng!D17={_str(ws, 'D17')!r}"
        assert abs(_num(ws, "G17") - 341.32) < 0.01, f"Tráng!G17={_num(ws, 'G17')}"

    def test_trang2_params(self):
        ws = self.wb["Tráng 2"]
        assert _str(ws, "D31") == "97 ± 5", f"Tráng 2!D31={_str(ws, 'D31')!r}"
        assert _str(ws, "D50") == "63 ± 1", f"Tráng 2!D50={_str(ws, 'D50')!r}"
        assert _num(ws, "M32") == 1040, f"Tráng 2!M32={_num(ws, 'M32')}"
        assert _str(ws, "M33") == "160 ± 3", f"Tráng 2!M33={_str(ws, 'M33')!r}"

    def test_dan2_params(self):
        ws = self.wb["Dán 2"]
        assert _str(ws, "D37") == "40 ±3", f"Dán 2!D37={_str(ws, 'D37')!r}"
        assert _str(ws, "D45") == "920 ± 5", f"Dán 2!D45={_str(ws, 'D45')!r}"

    def test_regression_values(self):
        assert _num(self.ws, "G18") == 328.44, f"G18={_num(self.ws, 'G18')}"
        assert _num(self.ws, "G29") == 5250.0, f"G29={_num(self.ws, 'G29')}"
        assert _num(self.ws, "M6") == 5000.0, f"M6={_num(self.ws, 'M6')}"
        assert _num(self.ws, "M14") == 0.041, f"M14={_num(self.ws, 'M14')}"

    def test_sheets_visible(self):
        # printed + liner order: nothing hidden except template-default sheets
        for name in ("In", "Tráng", "Dán", "Thổi", "May", "X", "In 2", "Tráng 2", "Dán 2"):
            assert _sheet_state(self.wb, name) == "visible", f"{name} should be visible"

    def test_cleanliness_no_stray_fill(self):
        all_cells = IN_HEADER + IN_INFO_M_KP + IN_THANH_PHAM
        _check_no_solid_fill(self.ws, all_cells, "In")

    def test_ink_kg_blank(self):
        for r in range(20, 24):
            v = _num(self.ws, f"G{r}")
            assert v is None or v == 0, f"In!G{r}={v} should be blank"

    def test_no_ink_rows(self):
        # Mực là vật tư điền thủ công theo thiết kế — sheet In không được còn tên/mã mực
        for r in range(20, 24):
            assert _str(self.ws, f"C{r}") == "", \
                f"In!C{r} không được chứa tên mực: {_str(self.ws, f'C{r}')!r}"
            assert _str(self.ws, f"D{r}") == "", \
                f"In!D{r} không được chứa mã mực: {_str(self.ws, f'D{r}')!r}"


class Test40KgOpp:
    """OPP 40kg 4 Oranges -- 3 mau in."""

    @classmethod
    def setup_class(cls):
        cls.family, cls.fields, cls.wb = _make_fixture(dict(SAMPLE_40KG), 3, "dm_test_40kg_")
        cls.ws = cls.wb["In"]

    @classmethod
    def teardown_class(cls):
        cls.wb.close()

    def test_completeness_header(self):
        _check_all_filled(self.ws, IN_HEADER, "In header")

    def test_completeness_info_m(self):
        _check_all_filled(self.ws, IN_INFO_M_OPP, "In info_M")

    def test_completeness_info_c(self):
        _check_all_filled(self.ws, IN_INFO_C, "In info_C")

    def test_completeness_thanh_pham(self):
        _check_all_filled(self.ws, IN_THANH_PHAM, "In thanh_pham")

    def test_opp_material_filled(self):
        _check_all_filled(self.ws, ["C18", "D18", "G18", "G19", "G20", "G21"], "In opp")

    def test_opp_material_name(self):
        assert "Màng BOPP mờ K1040" in _str(self.ws, "C18"), \
            f"In!C18 got {_str(self.ws, 'C18')!r}"
        assert _str(self.ws, "D18") == "MB01810400001", f"D18={_str(self.ws, 'D18')!r}"

    def test_stage_values_filled(self):
        for sheet, cells in {
            "Tráng": ["C17", "D17", "G17", "C18", "C19", "G18", "G19", "G20"],
            "Dán":   ["C17", "D17", "G17", "C18", "G18"],
            "Thổi":  ["C17", "G17", "G18"],
            "May":   ["G17", "G18", "C18"],
        }.items():
            _check_all_filled(self.wb[sheet], cells, f"{sheet} stage")

    def test_trang_manh(self):
        ws = self.wb["Tráng"]
        assert _str(ws, "D17") == "MW07510600001", f"Tráng!D17={_str(ws, 'D17')!r}"

    def test_in1_sizes(self):
        ws = self.wb["In 1"]
        assert _str(ws, "C17") == "420+80 x 820", f"In 1!C17={_str(ws, 'C17')!r}"
        assert _str(ws, "C18") == "420+80 x 820", f"In 1!C18={_str(ws, 'C18')!r}"

    def test_trang2_params(self):
        ws = self.wb["Tráng 2"]
        assert _str(ws, "D30") == "97 ± 5", f"Tráng 2!D30={_str(ws, 'D30')!r}"
        assert _str(ws, "D49") == "63 ± 1", f"Tráng 2!D49={_str(ws, 'D49')!r}"
        assert _num(ws, "M31") == 1040, f"Tráng 2!M31={_num(ws, 'M31')}"

    def test_dan2_params(self):
        ws = self.wb["Dán 2"]
        assert _str(ws, "D37") == "40 ±3", f"Dán 2!D37={_str(ws, 'D37')!r}"
        assert _str(ws, "D45") == "820 ± 5", f"Dán 2!D45={_str(ws, 'D45')!r}"

    def test_thoi2_filled(self):
        ws = self.wb["Thổi 2"]
        assert _str(ws, "D49") == "51 ± 1", f"Thổi 2!D49={_str(ws, 'D49')!r}"

    def test_trang2_4oranges_ratios(self):
        ws = self.wb["Tráng 2"]
        assert abs(_num(ws, "E22") - 0.07) < 1e-9, f"Tráng 2!E22={_num(ws, 'E22')}"
        assert abs(_num(ws, "E23") - 0.08) < 1e-9, f"Tráng 2!E23={_num(ws, 'E23')}"

    def test_thoi2_thuong_ratios(self):
        ws = self.wb["Thổi 2"]
        assert abs(_num(ws, "E18") - 0.893) < 1e-9, f"Thổi 2!E18={_num(ws, 'E18')}"
        assert abs(_num(ws, "E19") - 0.107) < 1e-9, f"Thổi 2!E19={_num(ws, 'E19')}"

    def test_regression_values(self):
        assert _num(self.ws, "M9") == 2983.0, f"M9={_num(self.ws, 'M9')}"
        assert abs(_num(self.ws, "G18") - 49.1407) < 0.01, f"G18={_num(self.ws, 'G18')}"
        assert abs(_num(self.ws, "G20") - 12.5286) < 0.01, f"G20={_num(self.ws, 'G20')}"
        assert abs(_num(self.ws, "G21") - 5.3694) < 0.01, f"G21={_num(self.ws, 'G21')}"
        assert _num(self.ws, "G29") == 3000.0, f"G29={_num(self.ws, 'G29')}"

    def test_dan_glue(self):
        ws = self.wb["Dán"]
        assert abs(_num(ws, "G17") - 4.0217) < 0.01, f"Dán!G17={_num(ws, 'G17')}"
        assert abs(_num(ws, "G18") - 4.0217) < 0.01, f"Dán!G18={_num(ws, 'G18')}"

    def test_cleanliness_no_stray_fill(self):
        all_cells = IN_HEADER + IN_INFO_M_OPP + IN_INFO_C + IN_THANH_PHAM
        _check_no_solid_fill(self.ws, all_cells, "In")

    def test_ink_kg_blank(self):
        for r in range(22, 29):
            v = _num(self.ws, f"G{r}")
            assert v is None or v == 0, f"In!G{r}={v} should be blank"

    def test_no_ink_rows(self):
        # Mực là vật tư điền thủ công theo thiết kế — sheet In không được còn tên/mã mực
        for r in range(22, 29):
            assert _str(self.ws, f"C{r}") == "", \
                f"In!C{r} không được chứa tên mực: {_str(self.ws, f'C{r}')!r}"
            assert _str(self.ws, f"D{r}") == "", \
                f"In!D{r} không được chứa mã mực: {_str(self.ws, f'D{r}')!r}"


class TestNormal40KgOpp:
    """Normal OPP order (Hãng Sơn Đông Á) with PE-rin liner, 2 màu in.

    Regression for the three fixed bugs:
      * Tráng taical 10% / hạt màu 5% (7/8 chỉ dành cho 4 Oranges & Thanh Phụng).
      * Thổi PE rin = 100% LDPE, không taical (tỷ lệ ở sheet Thổi 2).
      * Đóng gói OPP dài 82cm = 300 bao/kiện (bảng TSVH ĐÓNG GÓI).
    """

    @classmethod
    def setup_class(cls):
        cls.family, cls.fields, cls.wb = _make_fixture(dict(SAMPLE_40KG_NORMAL), 2, "dm_test_normal_opp_")
        cls.ws = cls.wb["In"]

    @classmethod
    def teardown_class(cls):
        cls.wb.close()

    def test_trang_ratios_normal_10_5(self):
        ws = self.wb["Tráng"]
        assert abs(_num(ws, "G19") - 9.1266) < 0.001, f"Tráng!G19={_num(ws, 'G19')}"
        assert abs(_num(ws, "G20") - 4.5633) < 0.001, f"Tráng!G20={_num(ws, 'G20')}"
        assert _str(ws, "C20") == "Hạt màu", f"Tráng!C20={_str(ws, 'C20')!r}"

    def test_trang2_ratio_cells_normal(self):
        ws = self.wb["Tráng 2"]
        assert abs(_num(ws, "E21") - 0.85) < 1e-9, f"Tráng 2!E21={_num(ws, 'E21')}"
        assert abs(_num(ws, "E22") - 0.10) < 1e-9, f"Tráng 2!E22={_num(ws, 'E22')}"
        assert abs(_num(ws, "E23") - 0.05) < 1e-9, f"Tráng 2!E23={_num(ws, 'E23')}"

    def test_thoi2_rin_100pct_ldpe(self):
        ws = self.wb["Thổi 2"]
        assert abs(_num(ws, "E18") - 1.0) < 1e-9, f"Thổi 2!E18={_num(ws, 'E18')}"
        assert abs(_num(ws, "E19") - 0.0) < 1e-9, f"Thổi 2!E19={_num(ws, 'E19')}"

    def test_thoi2_dims_tolerance(self):
        # túi lồng rin 50x92cm 20gr → dung sai: rộng/dài ±10, độ dày/khối lượng ±1
        ws = self.wb["Thổi 2"]
        assert _str(ws, "D45") == "500 ± 10", f"Thổi 2!D45={_str(ws, 'D45')!r}"
        assert _str(ws, "D46") == "920 ± 10", f"Thổi 2!D46={_str(ws, 'D46')!r}"
        assert _str(ws, "D48") == "23 ± 1", f"Thổi 2!D48={_str(ws, 'D48')!r}"
        assert _str(ws, "D49") == "20 ± 1", f"Thổi 2!D49={_str(ws, 'D49')!r}"

    def test_thoi_sheet_rin_quantities(self):
        ws = self.wb["Thổi"]
        assert abs(_num(ws, "G17") - 105.0) < 0.01, f"Thổi!G17={_num(ws, 'G17')}"
        assert _num(ws, "G18") == 0.0, f"Thổi!G18={_num(ws, 'G18')}"

    def test_dong_goi_300_bao_kien(self):
        ws = self.wb["Đóng gói 2"]
        assert _str(ws, "C32") == "300 bao/kiện", f"Đóng gói 2!C32={_str(ws, 'C32')!r}"
        assert self.fields.get("bao_kien") == 300, f"fields.bao_kien={self.fields.get('bao_kien')}"

    def test_dong_goi_header_filled_and_residue_cleared(self):
        ws = self.wb["Đóng gói 2"]
        assert _str(ws, "C5") == "26BA2IKH00010000004", f"Đóng gói 2!C5={_str(ws, 'C5')!r}"
        assert _str(ws, "C7") == "Bao BT Waler Nội Thất Cao Cấp 40kg", f"Đóng gói 2!C7={_str(ws, 'C7')!r}"
        assert _str(ws, "C16") == "5000", f"Đóng gói 2!C16={_str(ws, 'C16')!r}"
        for ref in ("V5", "W5", "X5"):
            assert _str(ws, ref) == "", f"Đóng gói 2!{ref} residue not cleared: {_str(ws, ref)!r}"


class TestKhongIn:
    """không in: In sheets hidden, X image cleared, main material on Tráng row 4."""

    @classmethod
    def setup_class(cls):
        order = dict(SAMPLE_25KG)
        order["has_print"] = False
        cls.family, cls.fields, cls.wb = _make_fixture(order, 0, "dm_test_khongin_")
        cls.ws = cls.wb["Tráng"]

    @classmethod
    def teardown_class(cls):
        cls.wb.close()

    def test_in_sheets_hidden(self):
        for name in ("In", "In 1", "In 2"):
            assert _sheet_state(self.wb, name) == "hidden", f"{name} should be hidden"

    def test_x_image_cleared(self):
        assert len(self.wb["X"]._images) == 0, "X sheet should have no image when không in"

    def test_main_material_on_trang_row4(self):
        assert _str(self.ws, "C20") == "Giấy Kraft vàng Nhật K1020 ĐL70 TAIKO", \
            f"Tráng!C20={_str(self.ws, 'C20')!r}"
        assert _str(self.ws, "D20") == "GN07010200001", f"D20={_str(self.ws, 'D20')!r}"
        assert abs(_num(self.ws, "G20") - 328.44) < 0.01, f"Tráng!G20={_num(self.ws, 'G20')}"

    def test_kho_mang_validation_skipped(self):
        # OPP without print must not require kho_mang
        from generate import validate_inputs
        order = dict(SAMPLE_40KG)
        order["has_print"] = False
        del order["kho_mang"]
        order["width_cm"] = 0
        order["gusset_cm"] = 0
        validate_inputs(order, "opp", 0)  # should not raise


class TestKhongLong:
    """không lồng túi: Thổi sheets hidden, May 2 quy cách dây blank."""

    @classmethod
    def setup_class(cls):
        order = dict(SAMPLE_40KG)
        order["has_tui_long"] = False
        cls.family, cls.fields, cls.wb = _make_fixture(order, 3, "dm_test_khonglong_")

    @classmethod
    def teardown_class(cls):
        cls.wb.close()

    def test_thoi_sheets_hidden(self):
        for name in ("Thổi", "Thổi 1", "Thổi 2"):
            assert _sheet_state(self.wb, name) == "hidden", f"{name} should be hidden"

    def test_may2_qc_blank(self):
        assert _str(self.wb["May 2"], "C37") == "", "May 2!C37 should be blank"


class TestInputParsing:
    """Tests for _parse_qty, _extract_pe_liner_weight, validate_inputs."""

    def test_parse_qty_none_empty(self):
        from generate import _parse_qty
        assert _parse_qty(None) is None
        assert _parse_qty("") is None

    def test_parse_qty_int_passthrough(self):
        from generate import _parse_qty
        assert _parse_qty(5000) == 5000
        assert _parse_qty(0) == 0

    def test_parse_qty_plain_int_str(self):
        from generate import _parse_qty
        assert _parse_qty("5000") == 5000

    def test_parse_qty_vn_thousands_dot(self):
        from generate import _parse_qty
        assert _parse_qty("5.000") == 5000

    def test_parse_qty_vn_thousands_comma(self):
        from generate import _parse_qty
        assert _parse_qty("5,000") == 5000

    def test_parse_qty_multiple_dots(self):
        from generate import _parse_qty
        assert _parse_qty("1.234.567") == 1234567

    def test_parse_qty_decimal_dot(self):
        from generate import _parse_qty
        assert _parse_qty("0.85") == 0.85

    def test_parse_qty_decimal_comma(self):
        from generate import _parse_qty
        assert _parse_qty("1,5") == 1.5

    def test_extract_pe_liner_4oranges_spec(self):
        from generate import _extract_pe_liner_weight
        spec = ("5. Quy cách lồng túi PE: Lồng túi PE thường 50gr 50x100cm (MTLS)")
        result = _extract_pe_liner_weight(spec)
        assert result == 0.05, f"Expected 0.05, got {result}"

    def test_extract_pe_liner_tui_long_50gram(self):
        from generate import _extract_pe_liner_weight
        result = _extract_pe_liner_weight("túi lồng 50gram")
        assert result == 0.05, f"Expected 0.05, got {result}"

    def test_extract_pe_liner_ignores_bag_capacity(self):
        from generate import _extract_pe_liner_weight
        assert _extract_pe_liner_weight("Bao KP 25kg") is None
        assert _extract_pe_liner_weight("40KG") is None

    def test_parse_spec_with_liner(self):
        from generate import parse_spec
        spec = ("1.Kích thước: (42+8) cm x 82cm\n"
                "5. Quy cách lồng túi PE: Lồng túi PE thường 50gr 50x100cm")
        out = parse_spec(spec)
        assert out.get("width_cm") == 42
        assert out.get("gusset_cm") == 8
        assert out.get("inner_bag_weight_kg") == 0.05

    def test_parse_spec_kp_no_liner(self):
        from generate import parse_spec
        out = parse_spec("55 x 85 cm")
        assert out.get("width_plus_gusset_m") == 0.55
        assert out.get("bag_length_m") == 0.85
        assert "inner_bag_weight_kg" not in out

    def test_validate_inputs_raises_on_zero_qty(self):
        from generate import validate_inputs, InputValidationError
        import copy
        order = copy.deepcopy(SAMPLE_25KG)
        order["qty"] = 0
        try:
            validate_inputs(order, "paper_kp", 2)
            assert False, "expected InputValidationError"
        except InputValidationError as e:
            assert "Số lượng" in str(e)

    def test_validate_inputs_accepts_zero_colors(self):
        from generate import validate_inputs
        # 0 = không in — valid even for a bag auto-detected as printed
        validate_inputs(SAMPLE_25KG, "paper_kp", 0)

    def test_validate_inputs_raises_on_negative_colors(self):
        from generate import validate_inputs, InputValidationError
        try:
            validate_inputs(SAMPLE_25KG, "paper_kp", -1)
            assert False, "expected InputValidationError"
        except InputValidationError as e:
            assert "không được âm" in str(e)

    def test_compute_zero_colors_forces_khong_in(self):
        from generate import compute
        fields = compute(dict(SAMPLE_25KG), "paper_kp", 0)
        assert fields["so_mau_in"] == 0, f"so_mau_in={fields['so_mau_in']}"
        assert fields["has_print"] is False, "0 màu must force không in"

    def test_validate_inputs_raises_on_missing_liner_ibw(self):
        from generate import validate_inputs, InputValidationError
        order = dict(SAMPLE_25KG)
        order["spec"] = "55 x 85 cm\nLồng túi PE thường 50gr"
        order["product_name"] = "Bao KP"
        del order["inner_bag_weight_kg"]
        try:
            validate_inputs(order, "paper_kp", 2)
            assert False, "expected InputValidationError"
        except InputValidationError as e:
            assert "túi lồng" in str(e) or "inner_bag_weight" in str(e)

    def test_validate_inputs_khong_long_tui_no_ibw_ok(self):
        # spec says "Quy cách lồng túi PE: không lồng túi" → must NOT require ibw
        from generate import validate_inputs
        order = dict(SAMPLE_25KG)
        order["spec"] = ("1.Kích thước: 45x75 cm\n"
                         "5. Quy cách lồng túi PE: không lồng túi")
        order["product_name"] = "Bao KP"
        del order["inner_bag_weight_kg"]
        validate_inputs(order, "paper_kp", 2)  # should not raise

    def test_has_tui_long_khong_long(self):
        from generate import _has_tui_long
        order = dict(SAMPLE_25KG)
        order["spec"] = "5. Quy cách lồng túi PE: không lồng túi"
        order["product_name"] = "Bao KP"
        assert _has_tui_long(order) is False

    def test_validate_inputs_opp_missing_kho_mang(self):
        from generate import validate_inputs, InputValidationError
        order = dict(SAMPLE_40KG)
        del order["kho_mang"]
        order["width_cm"] = 0
        order["gusset_cm"] = 0
        try:
            validate_inputs(order, "opp", 3)
            assert False, "expected InputValidationError"
        except InputValidationError as e:
            assert "kho_mang" in str(e)

    def test_validate_inputs_valid_sample_passes(self):
        from generate import validate_inputs
        # SAMPLE_25KG is paper_kp, colors=2, has valid ibw
        validate_inputs(SAMPLE_25KG, "paper_kp", 2)

    def test_ycsx_parse_4oranges(self):
        ycsx_path = os.path.join(HERE, "..", "samples", "YCSX_4oranges.xlsx")
        if not os.path.isfile(ycsx_path):
            import pytest
            pytest.skip("YCSX_4oranges.xlsx not found")
        from generate import parse_ycsx
        order = parse_ycsx(ycsx_path)
        assert order.get("customer") == "CÔNG TY 4 ORANGES CO.,LTD"
        products = order.get("products", [])
        assert len(products) >= 1
        # first product: qty=5000
        assert products[0].get("qty") == 5000, f"qty={products[0].get('qty')}"
        assert order.get("inner_bag_weight_kg") == 0.05, \
            f"ibw={order.get('inner_bag_weight_kg')}"


class TestYcsxPreserve:
    """YCSX output must be cloned from the input YCSX — mã code, mã khách hàng,
    và khối "chất lượng yêu cầu" các công đoạn (MÀNH/IN/TRÁNG/DÁN/THỔI/MAY) được
    giữ nguyên. Regression: trước đây fill_ycsx xóa sạch dòng 14–23 của template."""

    @classmethod
    def setup_class(cls):
        ycsx_path = os.path.join(HERE, "..", "samples", "YCSX_4oranges.xlsx")
        if not os.path.isfile(ycsx_path):
            import pytest
            pytest.skip("YCSX_4oranges.xlsx not found")
        outdir = tempfile.mkdtemp(prefix="ycsx_preserve_")
        run(("ycsx", ycsx_path), 3, outdir=outdir)
        ycsx_files = [os.path.join(outdir, f) for f in os.listdir(outdir)
                      if f.endswith(".xlsx") and f.startswith("YCSX")]
        assert ycsx_files, f"no YCSX output in {os.listdir(outdir)}"
        wb = load_workbook(ycsx_files[0], data_only=True)
        cls.ws = wb[wb.sheetnames[0]]
        wb.close()

    def test_customer_code_preserved(self):
        assert "Mã khách hàng: KH00565" in _str(self.ws, "B9"), \
            f"B9={_str(self.ws, 'B9')!r}"

    def test_ma_code_preserved(self):
        assert _str(self.ws, "E13") == "M4-13-2603-13", f"E13={_str(self.ws, 'E13')!r}"

    def test_stage_quality_block_preserved(self):
        stages = [_str(self.ws, f"C{r}") for r in range(16, 22)]
        assert stages == ["MÀNH", "IN", "TRÁNG", "DÁN", "THỔI", "MAY"], \
            f"stages={stages}"
        assert _str(self.ws, "D16") == "A : CLRC", f"D16={_str(self.ws, 'D16')!r}"

    def test_product_rows_refreshed(self):
        assert _str(self.ws, "C13") == "XD1KH00565BP0001", f"C13={_str(self.ws, 'C13')!r}"
        assert _str(self.ws, "J13") == "5000", f"J13={_str(self.ws, 'J13')!r}"
        assert _str(self.ws, "C14") == "XD1KH00565BP0004", f"C14={_str(self.ws, 'C14')!r}"

    def test_delivery_note_preserved(self):
        assert "Long An" in _str(self.ws, "B23"), f"B23={_str(self.ws, 'B23')!r}"

"""
Tra cuu Ten NVL + Ma NVL tu file "Nguyen vat lieu.xlsx" (danh muc chuan cua cong ty).

Dung thay cho viec tu dat ten/ma hoac giu nguyen ten/ma cu sot lai trong template.
Neu khong tim thay kho/dinh luong khop chinh xac, ham tra ve None kem danh sach
ung vien gan dung de Claude hoi lai nguoi dung, KHONG tu doan.

LUU Y QUAN TRONG: file nguon dung Unicode dang to hop (NFD, vi du "a" + dau huyen
rieng) thay vi dang dung san (NFC). Script nay luon normalize ve NFC truoc khi so
sanh - neu ban tu viet them ham tra cuu, nho lam tuong tu, neu khong so sanh chuoi
se am tham khong khop du nhin bang mat thi giong het nhau.
"""
import unicodedata
import openpyxl


def _norm(s):
    return unicodedata.normalize("NFC", s) if isinstance(s, str) else s


def load_nvl_list(nvl_file_path):
    """Doc toan bo danh muc NVL thanh list[dict(stt, ma, ten, dvt)], da normalize Unicode NFC."""
    wb = openpyxl.load_workbook(nvl_file_path, data_only=True)
    ws = wb["Sheet1"]
    items = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=4):
        stt, ma, ten, dvt = (c.value for c in row)
        if ma and ten:
            items.append({
                "stt": stt,
                "ma": _norm(str(ma).strip()),
                "ten": _norm(str(ten).strip()),
                "dvt": dvt,
            })
    return items


def find_by_kho(items, prefix_keywords, kho_mm, dinh_luong=None):
    """
    Tim vat lieu co ten chua prefix_keywords (list tu khoa bat buoc, khong phan biet hoa
    thuong) va kho dung bang kho_mm (dang "K<so>" trong ten). Neu truyen dinh_luong, uu tien
    ung vien co "DL<dinh_luong>" trong ten.

    Tra ve: (best_match_dict_or_None, list_ung_vien_gan_dung)
    """
    prefix_keywords = [_norm(k).lower() for k in prefix_keywords]
    kho_tag = f"k{int(round(kho_mm))}"
    candidates = []
    for it in items:
        ten_lower = it["ten"].lower()
        if all(kw in ten_lower for kw in prefix_keywords):
            candidates.append(it)

    exact = [c for c in candidates if kho_tag in c["ten"].lower()]
    if dinh_luong is not None:
        dl_tag = _norm(f"đl{int(dinh_luong)}").lower()
        exact_dl = [c for c in exact if dl_tag in c["ten"].lower()]
        if exact_dl:
            return exact_dl[0], exact_dl
    if exact:
        return exact[0], exact

    # Khong khop kho chinh xac -> tra ve gan dung de hoi lai, khong tu chon
    return None, candidates


def find_manh_trang(items, kho_mm, dinh_luong=70):
    """Mành trắng K<kho_mm> ĐL<dinh_luong>. VD kho_mm=1160, dinh_luong=70 -> K1160 ĐL70."""
    return find_by_kho(items, ["Mành", "trắng"], kho_mm, dinh_luong)


def find_giay_kraft(items, mau, xuatxu, kho_mm, dinh_luong=70, hang_uu_tien=None):
    """
    Giay Kraft <mau> <xuatxu> K<kho_mm> DL<dinh_luong> [<hang>].
    mau: "vàng" | "trắng"; xuatxu: "Nhật" | "Việt" | "Thái" | "Bãi Bằng" ...
    hang_uu_tien: VD "TAIKO" -> neu co nhieu hang cung kho, uu tien hang nay.
    """
    match, candidates = find_by_kho(items, ["Giấy Kraft", mau, xuatxu], kho_mm, dinh_luong)
    if match and hang_uu_tien and len(candidates) > 1:
        preferred = [c for c in candidates if _norm(hang_uu_tien).lower() in c["ten"].lower()]
        if preferred:
            return preferred[0], candidates
    return match, candidates


def find_exact_name(items, name_contains_all):
    """Tim theo danh sach tu khoa phai xuat hien het trong ten (khong can kho)."""
    name_contains_all = [_norm(k).lower() for k in name_contains_all]
    return [it for it in items if all(k in it["ten"].lower() for k in name_contains_all)]


def find_mang_opp(items, loai, kho_mm):
    """
    Mang BOPP <bong|mo> K<kho_mm> (VD "Mang BOPP mo K840 18 mic"). File goc dat ten
    dang "Mang BOPP <loai> K<kho> <so> mic" - khong loc theo mic (chi 1 dong/kho
    trong hau het truong hop), neu >1 ung vien cung kho se tra ve ung vien dau va
    danh sach du de doi chieu.
    """
    return find_by_kho(items, ["Màng BOPP", loai], kho_mm)


# Ten "cong doan" quy cach dong day tui long ma YCSX hay ghi (LTMS/MTLS) tuong
# ung voi loai day duoc dat ten trong Nguyen vat lieu.xlsx (chi co "đáy dài" /
# "đáy ngắn") - xac nhan boi nguoi dung ngay 2026-07-30:
#   LTMS -> đáy dài
#   MTLS -> đáy ngắn
QUY_CACH_DAY_TUI_LONG = {
    "ltms": "dài",
    "mtls": "ngắn",
    "long": "dài",       # phong khi YCSX ghi thang tieng Viet khong dau
    "dài": "dài",
    "ngắn": "ngắn",
    "ngan": "ngắn",
    # Xac nhan boi nguoi dung 2026-08-12: "may khong dinh day" = day ngan,
    # "may dinh day" = day dai.
    "may dính đáy": "dài",
    "dính đáy": "dài",
    "may không dính đáy": "ngắn",
    "không dính đáy": "ngắn",
    "khong dinh day": "ngắn",
}


def quy_cach_day_to_ten(quy_cach_raw):
    """Doi quy cach day tho tu YCSX (VD 'LTMS', 'MTLS', 'may khong dinh day')
    sang tag 'dài'/'ngắn' dung trong ten NVL. Tra ve None neu khong nhan dien
    duoc (khong doan).

    LUU Y: phai xet PHU DINH TRUOC - chuoi "không dính đáy" CHUA "dính đáy",
    neu tra dict theo thu tu tuy y se ra "dài" (nguoc hoan toan).
    """
    if not quy_cach_raw:
        return None
    key = _norm(str(quy_cach_raw)).strip().lower()
    if key in QUY_CACH_DAY_TUI_LONG:
        return QUY_CACH_DAY_TUI_LONG[key]
    if "không dính đáy" in key or "khong dinh day" in key:
        return "ngắn"
    if "dính đáy" in key or "dinh day" in key:
        return "dài"
    return None


def find_tui_long(items, loai, rong_cm, dai_cm, khoi_luong_g, day_loai=None):
    """
    Tim Tui long PE theo loai (thuong/rin) + rong x dai (cm) + khoi luong (gram),
    co the loc them theo day_loai ("dài"/"ngắn", suy tu quy_cach_day_to_ten()).
    Ten trong file khong co quy uoc chuan 1 kieu (VD "Túi PE thường 42x85cm 20gr,
    đáy dài", "Túi lồng PE rin 40x75, 20gr, đáy dài"...) nen so khop bang so
    (rong, dai, kg) xuat hien trong ten thay vi ghep chuoi cung nhau.

    Tra ve (best_match_or_None, candidates). Chi tra ve best_match khi CHI CO
    DUNG 1 ung vien khop ca 3 con so (rong/dai/kg) + dung loai (thuong/rin)
    [+ dung day_loai neu co truyen vao] - neu nhieu hon 1 hoac khong co, tra ve
    None kem candidates de hoi lai.
    """
    import re
    loai_norm = _norm(loai).lower()
    rong_i, dai_i, kg_i = int(round(rong_cm)), int(round(dai_cm)), int(round(khoi_luong_g))
    candidates = []
    for it in items:
        ten_lower = it["ten"].lower()
        if "túi" not in ten_lower or loai_norm not in ten_lower:
            continue
        nums = re.findall(r"\d+", it["ten"])
        nums_i = [int(n) for n in nums]
        if rong_i in nums_i and dai_i in nums_i and kg_i in nums_i:
            candidates.append(it)

    if day_loai:
        day_tag = _norm(f"đáy {day_loai}").lower()
        day_filtered = [c for c in candidates if day_tag in c["ten"].lower()]
        # QUAN TRONG (sua loi 2026-07-30, phat hien qua don Nong Nghiep Xanh): neu
        # da yeu cau day_loai cu the (VD LTMS -> "dai") thi CHI duoc tra ve ung
        # vien thuc su chua dung tag day do trong ten. Truoc day neu day_filtered
        # rong thi am tham fallback ve candidates chua loc (co the la dong "đáy
        # ngắn" trong khi YCSX yeu cau LTMS/"đáy dài") va tra ve nhu the la khop,
        # KHONG canh bao gi ca -> giao nham vat lieu ma khong ai biet. Gio neu
        # day_filtered rong, tra ve None (kem candidates cu de tham khao) de bat
        # buoc goi lai/hoi nguoi dung thay vi im lang chon sai.
        candidates_unfiltered_for_reference = candidates
        candidates = day_filtered
        if not candidates:
            return None, candidates_unfiltered_for_reference

    if len(candidates) == 1:
        return candidates[0], candidates
    # SUA LOI 2026-08-12 (don Hung Duy 25S03IKH00502000046): danh muc goc co
    # nhung dong NHAP TRUNG - cung mot ten, hai ma khac nhau (VD "Túi PE rin
    # 57x110cm 35gr, đáy ngắn" = TEA03505711001 va TEA0350571101). Truoc day
    # len(candidates) != 1 nen tra ve None -> sheet Thoi trong ten tui long VA
    # con sot ma cu cua template. Neu MOI ung vien cung mot ten thi thuc chat
    # chi la 1 vat lieu: tra ve dong dau. Caller thay len(candidates) > 1 thi
    # tu canh bao trung ma. Ten KHAC nhau that -> van tra None (khong doan).
    if candidates and len({_ten_key(c["ten"]) for c in candidates}) == 1:
        return candidates[0], candidates
    return None, candidates


def _ten_key(ten):
    """Khoa so sanh ten NVL: NFC + lower + bo moi khoang trang (nen "6 cm" ==
    "6cm", "52 x 110" == "52x110")."""
    return "".join(_norm(str(ten)).lower().split())


def _rong_cm_tags(rong_cm):
    """Cac cach danh muc goc viet chieu rong nep: 6 -> {"6cm"}; 6.5 ->
    {"6,5cm", "6.5cm"}."""
    if float(rong_cm) == int(rong_cm):
        return {f"{int(rong_cm)}cm"}
    txt = f"{float(rong_cm):g}"
    return {txt.replace(".", ",") + "cm", txt + "cm"}


def find_nep(items, tokens, rong_cm):
    """
    Tim Nep theo cac tu khoa KHONG PHU THUOC THU TU + chieu rong (cm).

    Vi sao khong dung find_exact_name() truc tiep: ten trong danh muc goc dat
    theo thu tu khac YCSX ("Nẹp KP vàng Nhật 6cm" vs YCSX ghi "Nẹp KP nhật
    vàng"), va chieu rong co the viet "6cm" hoac "6 cm" hoac "6,5cm". Ham nay
    so khop tren ten da bo khoang trang, va NEO chu so de "16cm" khong khop
    "6cm".

    tokens: list tu khoa bat buoc, VD ["nẹp kp", "vàng", "nhật"].
    Tra ve (best_match_or_None, candidates) - chi tra best_match khi CHI CO 1
    ung vien (hoac nhieu ung vien nhung trung ten y het), nguoc lai tra None
    kem candidates de hoi lai.
    """
    import re
    toks = [_ten_key(t) for t in tokens if t]
    tags = _rong_cm_tags(rong_cm)
    candidates = []
    for it in items:
        key = _ten_key(it["ten"])
        if not all(t in key for t in toks):
            continue
        if not any(re.search(r"(?<!\d)" + re.escape(tag), key) for tag in tags):
            continue
        candidates.append(it)
    if len(candidates) == 1:
        return candidates[0], candidates
    if candidates and len({_ten_key(c["ten"]) for c in candidates}) == 1:
        return candidates[0], candidates
    return None, candidates


# Cac vat lieu "co dinh" hay dung - tra 1 lan bang find_exact_name(items, [...]),
# dung lai cho moi don hang. Neu cong ty doi hang/ma, dung hardcode mu quang, hay tra lai.
COMMON_MATERIALS_HINT = {
    "F801C": ["Hạt nhựa nguyên sinh F801C"],
    "Taical (Tráng)": ["Hạt phụ gia EFPP 105BC"],
    "Vistamaxx": ["Hạt nhựa Vistamax 6202"],
    "LDPE": ["Hạt nhựa LLDPE", "FD21HN"],
    "Taical EFPE": ["Hạt taical trắng EFPE 75T"],
    "Chỉ may trắng": ["Chỉ may trắng"],
    "Dây bó bao": ["Dây bó bao"],
    "keo 9415 (Tân Châu/Neo Nam Việt)": ["Hạt nhựa PP FC9415"],
    "keo M9600": ["Hạt nhựa nguyên sinh M9600"],
    "Nẹp KP vàng Nhật 6cm": ["Nẹp KP vàng Nhật 6cm"],
    # LUU Y: danh muc goc viet SAI CHINH TA "Karft" (khong phai "Kraft") -
    # dung "kraft" lam tu khoa se KHONG khop dong nao.
    "Nẹp giấy Karft nhật vàng 10cm": ["nẹp giấy", "nhật", "vàng", "10cm"],
}


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Dung: python nvl_lookup.py <duong_dan_file_Nguyen_vat_lieu.xlsx>")
        sys.exit(1)
    items = load_nvl_list(sys.argv[1])
    print(f"Da doc {len(items)} dong NVL.")
    m, cands = find_manh_trang(items, 1160, 70)
    print("Mành trắng K1160 ĐL70:", m)
    g, cands = find_giay_kraft(items, "vàng", "Nhật", 1120, 70, hang_uu_tien="TAIKO")
    print("Giấy Kraft vàng Nhật K1120 ĐL70 (uu tien TAIKO):", g)
